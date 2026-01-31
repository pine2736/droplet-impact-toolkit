# droplet_diameter_tool_manual_timeline_overlay.py
# conda install -c conda-forge pyside6 pillow openpyxl numpy -y

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from PySide6.QtCore import Qt, QRectF, QPointF
from PySide6.QtGui import QAction, QKeySequence, QPainter, QPen, QImage, QPixmap
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QHBoxLayout, QVBoxLayout,
    QPushButton, QLabel, QComboBox, QLineEdit, QTableWidget, QTableWidgetItem,
    QMessageBox, QHeaderView, QSplitter, QGraphicsView, QGraphicsScene,
    QGraphicsPixmapItem, QGraphicsLineItem, QSizePolicy, QAbstractItemView, QMenu,
    QSlider, QGroupBox, QFormLayout, QDoubleSpinBox
)

from PIL import Image
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


SUPPORTED_EXTS = {".png", ".bmp", ".jpg", ".jpeg"}


def natural_sorted(paths: List[Path]) -> List[Path]:
    return sorted(paths, key=lambda p: p.name)


def parse_condition(folder_name: str) -> Tuple[str, str]:
    parts = folder_name.split("_")
    if len(parts) >= 2:
        return parts[0], parts[1]
    return "", ""


def clamp_sheet_name(name: str) -> str:
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    for b in bad:
        name = name.replace(b, "_")
    return name[:31] if len(name) > 31 else name


def pil_to_qpixmap(pil_img: Image.Image) -> QPixmap:
    if pil_img.mode != "RGBA":
        pil_img = pil_img.convert("RGBA")
    data = pil_img.tobytes("raw", "RGBA")
    qimg = QImage(data, pil_img.size[0], pil_img.size[1], QImage.Format_RGBA8888)
    return QPixmap.fromImage(qimg)


@dataclass
class MeasurementRow:
    condition_folder: str
    concentration: str
    speed: str
    image_name: str
    image_path: str
    frame_index: int
    fps: float
    t0_index: Optional[int]
    t0_image_name: Optional[str]
    time_s: Optional[float]
    time_ms: Optional[float]
    needle_px: Optional[float] = None
    mm_per_px: Optional[float] = None
    dh_mm: Optional[float] = None
    dv_mm: Optional[float] = None
    d0_mm: Optional[float] = None


class ImageView(QGraphicsView):
    """
    - Ctrl + Wheel：缩放
    - Ctrl + Left Drag：平移
    - 普通 Left Click：采点
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setRenderHint(QPainter.Antialiasing)
        self.setMouseTracking(True)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorViewCenter)
        self.setDragMode(QGraphicsView.NoDrag)
        self.setFocusPolicy(Qt.StrongFocus)

        self._click_callback = None
        self._panning = False
        self._zoom_changed_callback = None

    def set_click_callback(self, cb):
        self._click_callback = cb

    def set_zoom_changed_callback(self, cb):
        self._zoom_changed_callback = cb

    def wheelEvent(self, event):
        if event.modifiers() & Qt.ControlModifier:
            angle = event.angleDelta().y()
            if angle == 0:
                return
            factor = 1.25 if angle > 0 else 0.8
            self.scale(factor, factor)
            if self._zoom_changed_callback:
                self._zoom_changed_callback()
            event.accept()
            return
        event.ignore()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and (event.modifiers() & Qt.ControlModifier):
            self._panning = True
            self.setDragMode(QGraphicsView.ScrollHandDrag)
            super().mousePressEvent(event)
            return

        if event.button() == Qt.LeftButton and self._click_callback is not None:
            pos = self.mapToScene(event.position().toPoint())
            self._click_callback(pos)

        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self._panning and event.button() == Qt.LeftButton:
            self._panning = False
            self.setDragMode(QGraphicsView.NoDrag)
        super().mouseReleaseEvent(event)


class MainWindow(QMainWindow):
    """
    操作逻辑（按你的习惯）：
    - 当前图片可显示 3 条线：CAL/DH/DV
    - 只要“重新开始测量”（点击 CAL/DH/DV 任意按钮）：
        -> 立刻清掉当前图片上的三条线（全部清空）
        -> 开始新的两点测量
    - 两点完成后：画细线保留在图上，用于复核
    - 切换图片：若该图片已有线，自动重建显示
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Droplet Diameter Measure (Manual + Timeline + Overlay)")
        self.resize(1450, 900)

        # ---- Data ----
        self.root_dir: Optional[Path] = None
        self.condition_folders: List[Path] = []
        self.current_condition: Optional[Path] = None
        self.current_images: List[Path] = []
        self.current_image_index: int = -1

        self.calibration_mm_per_px: Dict[str, float] = {}
        self.calibration_needle_px: Dict[str, float] = {}

        # per-condition T0
        self.t0_index_by_cond: Dict[str, int] = {}
        self.t0_name_by_cond: Dict[str, str] = {}

        # results keyed by image_path
        self.results: Dict[str, MeasurementRow] = {}

        # current temporary measurement
        self.temp_dh_mm: Optional[float] = None
        self.temp_dv_mm: Optional[float] = None

        # measurement mode
        self.mode: Optional[str] = None  # CAL / DH / DV
        self.first_point: Optional[QPointF] = None

        # guide items (temporary)
        self.guide_items: List[object] = []

        # overlay lines per image (persistent for visual checking)
        # { image_path_str: { "CAL": QGraphicsLineItem, "DH": QGraphicsLineItem, "DV": QGraphicsLineItem } }
        self.overlay_by_image: Dict[str, Dict[str, QGraphicsLineItem]] = {}

        self.auto_fit_on_resize: bool = True

        # ---- UI ----
        central = QWidget()
        self.setCentralWidget(central)
        splitter = QSplitter(Qt.Horizontal, central)

        self.scene = QGraphicsScene(self)
        self.view = ImageView(self)
        self.view.setScene(self.scene)
        self.view.set_click_callback(self.on_scene_click)
        self.view.set_zoom_changed_callback(self.on_user_zoomed)
        self.view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        right.setMinimumWidth(520)
        right.setMaximumWidth(740)

        # root selection
        row1 = QHBoxLayout()
        self.root_edit = QLineEdit()
        self.root_edit.setPlaceholderText(r'Input folder path, e.g. D:\data\...')
        self.btn_browse = QPushButton("Browse...")
        self.btn_load = QPushButton("Load")
        row1.addWidget(self.root_edit, 1)
        row1.addWidget(self.btn_browse)
        row1.addWidget(self.btn_load)
        right_layout.addLayout(row1)

        # condition selector
        row2 = QHBoxLayout()
        self.condition_combo = QComboBox()
        self.btn_prev_cond = QPushButton("Prev Condition")
        self.btn_next_cond = QPushButton("Next Condition")
        row2.addWidget(QLabel("Condition:"))
        row2.addWidget(self.condition_combo, 1)
        row2.addWidget(self.btn_prev_cond)
        row2.addWidget(self.btn_next_cond)
        right_layout.addLayout(row2)

        # image navigation
        row3 = QHBoxLayout()
        self.btn_prev = QPushButton("Prev Image")
        self.btn_next = QPushButton("Next Image")
        row3.addWidget(self.btn_prev)
        row3.addWidget(self.btn_next)
        right_layout.addLayout(row3)

        # quick seek slider
        group_seek = QGroupBox("Quick Seek")
        seek_layout = QVBoxLayout(group_seek)
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setMinimum(0)
        self.slider.setMaximum(0)
        self.slider.setValue(0)
        self.slider.setSingleStep(1)
        self.slider.setPageStep(10)
        self.lbl_seek = QLabel("Frame: - / -")
        seek_layout.addWidget(self.slider)
        seek_layout.addWidget(self.lbl_seek)
        right_layout.addWidget(group_seek)

        self.lbl_img = QLabel("Image: -")
        self.lbl_img.setWordWrap(True)
        right_layout.addWidget(self.lbl_img)

        # fit
        row_fit = QHBoxLayout()
        self.btn_fit = QPushButton("Fit (Reset View)")
        row_fit.addWidget(self.btn_fit)
        right_layout.addLayout(row_fit)

        # needle
        row4 = QHBoxLayout()
        self.needle_edit = QLineEdit("0.41")
        self.needle_edit.setFixedWidth(90)
        row4.addWidget(QLabel("Needle OD (mm):"))
        row4.addWidget(self.needle_edit)
        row4.addStretch(1)
        right_layout.addLayout(row4)

        # timeline
        group_time = QGroupBox("Timeline")
        form = QFormLayout(group_time)

        self.fps_spin = QDoubleSpinBox()
        self.fps_spin.setRange(0.01, 100000.0)
        self.fps_spin.setDecimals(4)
        self.fps_spin.setValue(4000.0)
        self.fps_spin.setToolTip("Frames per second. Time = (frame_index - T0_index) / FPS")
        form.addRow("FPS:", self.fps_spin)

        self.btn_set_t0 = QPushButton("Set current image as T0 (t=0)")
        form.addRow(self.btn_set_t0)

        self.lbl_t0 = QLabel("T0: (not set)")
        self.lbl_time = QLabel("Time: (set T0 to enable)")
        self.lbl_time.setWordWrap(True)
        form.addRow(self.lbl_t0)
        form.addRow(self.lbl_time)

        right_layout.addWidget(group_time)

        # modes
        self.btn_cal = QPushButton("1) Calibrate Needle (Horizontal lock)")
        self.btn_dh = QPushButton("2) Measure Dh (Horizontal lock)")
        self.btn_dv = QPushButton("3) Measure Dv (Vertical lock)")
        right_layout.addWidget(self.btn_cal)
        right_layout.addWidget(self.btn_dh)
        right_layout.addWidget(self.btn_dv)

        # undo/reset
        row5 = QHBoxLayout()
        self.btn_undo_point = QPushButton("Undo Point (Ctrl+Z)")
        self.btn_reset_mode = QPushButton("Reset Step (Esc)")
        row5.addWidget(self.btn_undo_point)
        row5.addWidget(self.btn_reset_mode)
        right_layout.addLayout(row5)

        # status
        self.lbl_status = QLabel("Status: Load folder to start.")
        self.lbl_status.setWordWrap(True)
        self.lbl_cal = QLabel("Calibration: -")
        self.lbl_cal.setWordWrap(True)
        self.lbl_tmp = QLabel("Current: Dh=-  Dv=-  D0=-")
        self.lbl_tmp.setWordWrap(True)
        right_layout.addWidget(self.lbl_status)
        right_layout.addWidget(self.lbl_cal)
        right_layout.addWidget(self.lbl_tmp)

        self.lbl_hint = QLabel(
            "Hint:\n"
            "- Ctrl+Wheel=Zoom, Ctrl+Left Drag=Pan\n"
            "- CAL/DH: 2 clicks, horizontal locked\n"
            "- DV: 2 clicks, vertical locked\n"
            "- Click CAL/DH/DV to start: it will CLEAR ALL 3 lines on current image\n"
            "- Use slider to jump frames quickly\n"
            "- Set T0 then any frame shows relative time\n"
        )
        self.lbl_hint.setWordWrap(True)
        right_layout.addWidget(self.lbl_hint)

        # table
        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels([
            "Concentration", "Speed", "ImageName",
            "Frame", "Time(s)", "Dh(mm)", "Dv(mm)", "D0(mm)", "T0"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_table_context_menu)

        right_layout.addWidget(self.table, 1)

        # delete/export
        row6 = QHBoxLayout()
        self.btn_delete_row = QPushButton("Delete This Image Result")
        self.btn_export = QPushButton("Export Excel")
        row6.addWidget(self.btn_delete_row)
        row6.addWidget(self.btn_export)
        right_layout.addLayout(row6)

        splitter.addWidget(self.view)
        splitter.addWidget(right)
        splitter.setSizes([950, 500])
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 1)

        layout = QHBoxLayout(central)
        layout.addWidget(splitter)

        # ---- Shortcuts ----
        undo_act = QAction(self)
        undo_act.setShortcut(QKeySequence("Ctrl+Z"))
        undo_act.triggered.connect(self.undo_point)
        self.addAction(undo_act)

        reset_act = QAction(self)
        reset_act.setShortcut(QKeySequence("Esc"))
        reset_act.triggered.connect(self.reset_step)
        self.addAction(reset_act)

        select_all_act = QAction(self)
        select_all_act.setShortcut(QKeySequence("Ctrl+A"))
        select_all_act.triggered.connect(self.select_all_table)
        self.addAction(select_all_act)

        delete_act = QAction(self)
        delete_act.setShortcut(QKeySequence(Qt.Key_Delete))
        delete_act.triggered.connect(self.delete_selected_table_rows)
        self.addAction(delete_act)

        # ---- Signals ----
        self.btn_browse.clicked.connect(self.browse_root)
        self.btn_load.clicked.connect(self.load_root)

        self.condition_combo.currentIndexChanged.connect(self.on_condition_changed)
        self.btn_prev_cond.clicked.connect(lambda: self.step_condition(-1))
        self.btn_next_cond.clicked.connect(lambda: self.step_condition(1))

        self.btn_prev.clicked.connect(lambda: self.step_image(-1))
        self.btn_next.clicked.connect(lambda: self.step_image(1))
        self.slider.valueChanged.connect(self.on_slider_changed)

        self.btn_fit.clicked.connect(self.set_fit)

        # Start modes (will clear 3 lines for current image)
        self.btn_cal.clicked.connect(lambda: self.start_mode("CAL"))
        self.btn_dh.clicked.connect(lambda: self.start_mode("DH"))
        self.btn_dv.clicked.connect(lambda: self.start_mode("DV"))

        self.btn_undo_point.clicked.connect(self.undo_point)
        self.btn_reset_mode.clicked.connect(self.reset_step)

        self.btn_delete_row.clicked.connect(self.delete_current_image_result)
        self.btn_export.clicked.connect(self.export_excel)

        self.btn_set_t0.clicked.connect(self.set_current_as_t0)
        self.fps_spin.valueChanged.connect(self.update_time_labels)

        self.set_controls_enabled(False)

    # ---------------- UI helpers ----------------
    def set_controls_enabled(self, enabled: bool):
        for w in [
            self.condition_combo, self.btn_prev_cond, self.btn_next_cond,
            self.btn_prev, self.btn_next, self.btn_fit,
            self.btn_cal, self.btn_dh, self.btn_dv,
            self.btn_undo_point, self.btn_reset_mode,
            self.btn_delete_row, self.btn_export,
            self.table, self.slider,
            self.btn_set_t0, self.fps_spin
        ]:
            w.setEnabled(enabled)

    def set_status(self, text: str):
        self.lbl_status.setText(f"Status: {text}")

    def message(self, title: str, text: str, icon=QMessageBox.Information):
        m = QMessageBox(self)
        m.setIcon(icon)
        m.setWindowTitle(title)
        m.setText(text)
        m.exec()

    def on_user_zoomed(self):
        self.auto_fit_on_resize = False

    def set_fit(self):
        if self.scene.sceneRect().isNull():
            return
        self.view.resetTransform()
        self.view.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)
        self.auto_fit_on_resize = True

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.auto_fit_on_resize:
            self.set_fit()

    # ---------------- Folder loading ----------------
    def browse_root(self):
        d = QFileDialog.getExistingDirectory(self, "Select Root Folder", str(self.root_dir or Path.cwd()))
        if d:
            self.root_edit.setText(d)

    def collect_images(self, folder: Path) -> List[Path]:
        imgs = []
        for root, _, files in os.walk(folder):
            for fn in files:
                ext = Path(fn).suffix.lower()
                if ext in SUPPORTED_EXTS:
                    imgs.append(Path(root) / fn)
        return natural_sorted(imgs)

    def load_root(self):
        raw = self.root_edit.text().strip().strip('"')
        if not raw:
            self.message("Error", "Please input a folder path.")
            return
        p = Path(raw)
        if not p.exists() or not p.is_dir():
            self.message("Error", "Folder does not exist.")
            return

        self.root_dir = p
        candidates = []
        for sub in sorted([x for x in p.iterdir() if x.is_dir()], key=lambda x: x.name):
            imgs = self.collect_images(sub)
            if imgs:
                candidates.append(sub)

        self.condition_folders = candidates
        self.condition_combo.blockSignals(True)
        self.condition_combo.clear()
        for c in self.condition_folders:
            self.condition_combo.addItem(c.name)
        self.condition_combo.blockSignals(False)

        if not self.condition_folders:
            self.set_controls_enabled(False)
            self.set_status("No images found under subfolders.")
            return

        self.set_controls_enabled(True)
        self.set_status(f"Loaded {len(self.condition_folders)} conditions. Select one to start.")
        self.condition_combo.setCurrentIndex(0)

    def on_condition_changed(self, idx: int):
        if idx < 0 or idx >= len(self.condition_folders):
            return
        self.current_condition = self.condition_folders[idx]
        self.current_images = self.collect_images(self.current_condition)
        self.current_image_index = 0 if self.current_images else -1

        self.temp_dh_mm = None
        self.temp_dv_mm = None
        self.reset_step(clear_only_state=True)

        # update slider range
        self.slider.blockSignals(True)
        if self.current_images:
            self.slider.setMinimum(0)
            self.slider.setMaximum(len(self.current_images) - 1)
            self.slider.setValue(self.current_image_index)
        else:
            self.slider.setMinimum(0)
            self.slider.setMaximum(0)
            self.slider.setValue(0)
        self.slider.blockSignals(False)

        self.update_calibration_label()
        self.load_current_image()
        self.refresh_table()
        self.update_time_labels()

    def step_condition(self, delta: int):
        if not self.condition_folders:
            return
        i = self.condition_combo.currentIndex()
        ni = max(0, min(len(self.condition_folders) - 1, i + delta))
        self.condition_combo.setCurrentIndex(ni)

    def on_slider_changed(self, value: int):
        if not self.current_images:
            return
        value = max(0, min(len(self.current_images) - 1, int(value)))
        if value != self.current_image_index:
            self.current_image_index = value
            self.temp_dh_mm = None
            self.temp_dv_mm = None
            self.reset_step(clear_only_state=True)
            self.load_current_image()
            self.refresh_table()
            self.update_time_labels()

    def step_image(self, delta: int):
        if not self.current_images:
            return
        ni = self.current_image_index + delta
        ni = max(0, min(len(self.current_images) - 1, ni))
        if ni != self.current_image_index:
            self.current_image_index = ni
            self.slider.blockSignals(True)
            self.slider.setValue(ni)
            self.slider.blockSignals(False)

            self.temp_dh_mm = None
            self.temp_dv_mm = None
            self.reset_step(clear_only_state=True)
            self.load_current_image()
            self.refresh_table()
            self.update_time_labels()

    def load_current_image(self):
        self.scene.clear()
        self.guide_items.clear()
        self.auto_fit_on_resize = True

        if self.current_image_index < 0 or self.current_image_index >= len(self.current_images):
            self.lbl_img.setText("Image: -")
            self.lbl_seek.setText("Frame: - / -")
            return

        img_path = self.current_images[self.current_image_index]
        self.lbl_img.setText(f"Image: {img_path.name}   ({self.current_image_index + 1}/{len(self.current_images)})")
        self.lbl_seek.setText(f"Frame: {self.current_image_index} / {len(self.current_images) - 1}")

        try:
            pil_img = Image.open(img_path)
            pix = pil_to_qpixmap(pil_img)
        except Exception as e:
            self.message("Error", f"Failed to load image:\n{e}", icon=QMessageBox.Critical)
            return

        item = QGraphicsPixmapItem(pix)
        self.scene.addItem(item)
        self.scene.setSceneRect(QRectF(pix.rect()))

        self.set_fit()
        self.update_temp_label_from_saved()

        # IMPORTANT: rebuild overlay lines for this image (if any)
        self.rebuild_overlay_for_current_image()

        self.view.setFocus()

    # ---------------- Overlay: clear / draw / rebuild ----------------
    def clear_overlay_for_current_image(self):
        """
        清线：当前图片三条线（CAL/DH/DV）全部清掉。
        你要求：重新开始测量时，三条线必须都要清掉。
        """
        if self.current_image_index < 0 or not self.current_images:
            return
        img_path = str(self.current_images[self.current_image_index])
        m = self.overlay_by_image.get(img_path)
        if not m:
            return
        for _, item in list(m.items()):
            try:
                self.scene.removeItem(item)
            except Exception:
                pass
        self.overlay_by_image.pop(img_path, None)

    def draw_overlay_line(self, kind: str, p1: QPointF, p2: QPointF):
        """
        划线：在当前图片上画细线并保留（用于复核）。
        kind: "CAL" / "DH" / "DV"
        """
        if self.current_image_index < 0 or not self.current_images:
            return
        img_path = str(self.current_images[self.current_image_index])
        m = self.overlay_by_image.setdefault(img_path, {})

        # 如果同类线存在，先删掉（一般不会发生，因为 start_mode 已经清三条线）
        old = m.get(kind)
        if old is not None:
            try:
                self.scene.removeItem(old)
            except Exception:
                pass

        # 线尽量细
        pen = QPen(Qt.yellow if kind in ("CAL", "DH") else Qt.cyan)
        pen.setWidth(1)

        line = QGraphicsLineItem(p1.x(), p1.y(), p2.x(), p2.y())
        line.setPen(pen)
        line.setZValue(10)
        self.scene.addItem(line)

        m[kind] = line

    def rebuild_overlay_for_current_image(self):
        """
        重建线：切换图片时，把已保存的线重新加到 scene 里（用于检查）。
        注意：Qt item 只能属于一个 scene；我们是每次 load_current_image 先 scene.clear，
        所以 item 需要重新 addItem 才能显示。
        """
        if self.current_image_index < 0 or not self.current_images:
            return
        img_path = str(self.current_images[self.current_image_index])
        m = self.overlay_by_image.get(img_path)
        if not m:
            return

        # re-add the existing line items
        for item in m.values():
            try:
                # after scene.clear(), item's scene() is None
                if item.scene() is None:
                    self.scene.addItem(item)
            except Exception:
                # ignore if Qt complains
                pass

    # ---------------- Timeline / T0 ----------------
    def set_current_as_t0(self):
        if self.current_condition is None or not self.current_images or self.current_image_index < 0:
            return
        cond = self.current_condition.name
        self.t0_index_by_cond[cond] = self.current_image_index
        self.t0_name_by_cond[cond] = self.current_images[self.current_image_index].name
        self.set_status(f"Set T0 for [{cond}] at frame {self.current_image_index}: {self.t0_name_by_cond[cond]}")
        self.update_time_labels()
        self.refresh_table()

    def get_current_t0(self) -> Tuple[Optional[int], Optional[str]]:
        if self.current_condition is None:
            return None, None
        cond = self.current_condition.name
        return self.t0_index_by_cond.get(cond), self.t0_name_by_cond.get(cond)

    def compute_time_for_index(self, frame_index: int) -> Tuple[Optional[float], Optional[float]]:
        t0_idx, _ = self.get_current_t0()
        if t0_idx is None:
            return None, None
        fps = float(self.fps_spin.value())
        if fps <= 0:
            return None, None
        dt = (frame_index - t0_idx) / fps
        return dt, dt * 1000.0

    def update_time_labels(self):
        if self.current_condition is None or not self.current_images or self.current_image_index < 0:
            self.lbl_t0.setText("T0: (not set)")
            self.lbl_time.setText("Time: -")
            return

        t0_idx, t0_name = self.get_current_t0()
        if t0_idx is None:
            self.lbl_t0.setText("T0: (not set)")
            self.lbl_time.setText("Time: (set T0 to enable)")
            return

        self.lbl_t0.setText(f"T0: frame {t0_idx} | {t0_name}")

        dt_s, dt_ms = self.compute_time_for_index(self.current_image_index)
        fps = float(self.fps_spin.value())
        if dt_s is None:
            self.lbl_time.setText("Time: (invalid FPS or T0)")
            return

        interval_ms = 1000.0 / fps
        self.lbl_time.setText(
            f"Time: {dt_s:.6f} s  ({dt_ms:.3f} ms)\n"
            f"Frame interval: {interval_ms:.6f} ms @ {fps:.4f} FPS"
        )

    # ---------------- Measurement ----------------
    def start_mode(self, mode: str):
        """
        你要求的逻辑：重新开始测量时，三条线必须都清掉。
        所以这里无论 CAL/DH/DV，先 clear_overlay_for_current_image()。
        """
        if self.current_condition is None or not self.current_images:
            self.message("Hint", "Please load images first.")
            return

        # Clear ALL 3 lines on current image whenever user starts a measurement
        self.clear_overlay_for_current_image()

        if mode in ("DH", "DV") and self.current_condition.name not in self.calibration_mm_per_px:
            self.message("Need Calibration", "This condition is not calibrated yet.\nPlease calibrate needle first.")
            return

        self.mode = mode
        self.first_point = None
        self.clear_guides()

        if mode == "CAL":
            self.set_status("Calibration: click first point, then second point (horizontal locked).")
        elif mode == "DH":
            self.set_status("Measure Dh: click left edge, then right edge (horizontal locked).")
        elif mode == "DV":
            self.set_status("Measure Dv: click one edge, then the other (vertical locked).")

    def on_scene_click(self, pos: QPointF):
        if self.mode is None:
            return
        rect = self.scene.sceneRect()
        if rect.isNull() or not rect.contains(pos):
            return

        if self.first_point is None:
            self.first_point = pos
            self.draw_first_point_marker(pos)
            self.draw_guide_line(pos, self.mode)
        else:
            p1 = self.first_point
            p2 = pos

            if self.mode in ("CAL", "DH"):
                # horizontal locked
                p2 = QPointF(p2.x(), p1.y())
                px = abs(p2.x() - p1.x())
                if px < 1e-6:
                    self.message("Invalid", "Distance too small. Please reselect.")
                    self.reset_step()
                    return

                if self.mode == "CAL":
                    self.apply_calibration(px)
                    # keep line for visual check
                    self.draw_overlay_line("CAL", p1, p2)
                else:
                    self.apply_dh(px)
                    self.draw_overlay_line("DH", p1, p2)

            elif self.mode == "DV":
                # vertical locked
                p2 = QPointF(p1.x(), p2.y())
                px = abs(p2.y() - p1.y())
                if px < 1e-6:
                    self.message("Invalid", "Distance too small. Please reselect.")
                    self.reset_step()
                    return

                self.apply_dv(px)
                self.draw_overlay_line("DV", p1, p2)

            self.mode = None
            self.first_point = None
            self.clear_guides()

    def apply_calibration(self, needle_px: float):
        if self.current_condition is None:
            return
        cond_name = self.current_condition.name

        try:
            needle_mm = float(self.needle_edit.text().strip())
            if needle_mm <= 0:
                raise ValueError
        except Exception:
            self.message("Error", "Needle OD must be a positive number.")
            return

        mm_per_px = needle_mm / needle_px
        self.calibration_mm_per_px[cond_name] = mm_per_px
        self.calibration_needle_px[cond_name] = needle_px

        self.update_calibration_label()
        self.set_status(f"Calibrated: needle_px={needle_px:.2f}, mm_per_px={mm_per_px:.8f} (Condition: {cond_name})")
        self.update_temp_label()

    def get_mm_per_px_current(self) -> Optional[float]:
        if self.current_condition is None:
            return None
        return self.calibration_mm_per_px.get(self.current_condition.name)

    def apply_dh(self, dh_px: float):
        mm_per_px = self.get_mm_per_px_current()
        if mm_per_px is None:
            return
        self.temp_dh_mm = dh_px * mm_per_px
        self.set_status(f"Dh measured: {self.temp_dh_mm:.5f} mm")
        self.update_temp_label()
        self.try_save_if_complete()

    def apply_dv(self, dv_px: float):
        mm_per_px = self.get_mm_per_px_current()
        if mm_per_px is None:
            return
        self.temp_dv_mm = dv_px * mm_per_px
        self.set_status(f"Dv measured: {self.temp_dv_mm:.5f} mm")
        self.update_temp_label()
        self.try_save_if_complete()

    def try_save_if_complete(self):
        if self.temp_dh_mm is None or self.temp_dv_mm is None:
            return
        if self.current_condition is None or self.current_image_index < 0:
            return

        img_path = self.current_images[self.current_image_index]
        cond_name = self.current_condition.name
        conc, spd = parse_condition(cond_name)

        mm_per_px = self.calibration_mm_per_px.get(cond_name)
        needle_px = self.calibration_needle_px.get(cond_name)
        d0 = (self.temp_dh_mm ** 2 * self.temp_dv_mm) ** (1.0 / 3.0)

        t0_idx, t0_name = self.get_current_t0()
        dt_s, dt_ms = self.compute_time_for_index(self.current_image_index)
        fps = float(self.fps_spin.value())

        self.results[str(img_path)] = MeasurementRow(
            condition_folder=cond_name,
            concentration=conc,
            speed=spd,
            image_name=img_path.name,
            image_path=str(img_path),
            frame_index=self.current_image_index,
            fps=fps,
            t0_index=t0_idx,
            t0_image_name=t0_name,
            time_s=dt_s,
            time_ms=dt_ms,
            needle_px=needle_px,
            mm_per_px=mm_per_px,
            dh_mm=self.temp_dh_mm,
            dv_mm=self.temp_dv_mm,
            d0_mm=d0
        )
        self.refresh_table()
        self.set_status("Saved result for current image (overwrite if existed).")

    # ---------------- Guides & Undo ----------------
    def clear_guides(self):
        for it in self.guide_items:
            try:
                self.scene.removeItem(it)
            except Exception:
                pass
        self.guide_items.clear()

    def draw_guide_line(self, p1: QPointF, mode: str):
        pen = QPen(Qt.green)
        pen.setWidth(1)
        rect = self.scene.sceneRect()

        if mode in ("CAL", "DH"):
            y = p1.y()
            line = QGraphicsLineItem(rect.left(), y, rect.right(), y)
            line.setPen(pen)
            self.scene.addItem(line)
            self.guide_items.append(line)
        elif mode == "DV":
            x = p1.x()
            line = QGraphicsLineItem(x, rect.top(), x, rect.bottom())
            line.setPen(pen)
            self.scene.addItem(line)
            self.guide_items.append(line)

    def draw_first_point_marker(self, p: QPointF):
        pen = QPen(Qt.red)
        pen.setWidth(1)  # marker also thin
        size = 6
        l1 = QGraphicsLineItem(p.x() - size, p.y(), p.x() + size, p.y())
        l2 = QGraphicsLineItem(p.x(), p.y() - size, p.x(), p.y() + size)
        l1.setPen(pen)
        l2.setPen(pen)
        self.scene.addItem(l1)
        self.scene.addItem(l2)
        self.guide_items.extend([l1, l2])

    def undo_point(self):
        if self.mode is None:
            self.set_status("Nothing to undo (not in a measurement step).")
            return
        if self.first_point is None:
            self.set_status("Nothing to undo (no point selected yet).")
            return
        self.first_point = None
        self.clear_guides()
        self.set_status("Undid first point. Click again to start this step.")

    def reset_step(self, clear_only_state: bool = False):
        self.mode = None
        self.first_point = None
        self.clear_guides()
        if not clear_only_state:
            self.set_status("Step reset. Choose a measurement button to continue.")

    # ---------------- Table & labels ----------------
    def update_calibration_label(self):
        if self.current_condition is None:
            self.lbl_cal.setText("Calibration: -")
            return
        cond_name = self.current_condition.name
        if cond_name in self.calibration_mm_per_px:
            mpp = self.calibration_mm_per_px[cond_name]
            npx = self.calibration_needle_px.get(cond_name, None)
            if npx is None:
                self.lbl_cal.setText(f"Calibration: mm_per_px={mpp:.8f}")
            else:
                self.lbl_cal.setText(f"Calibration: needle_px={npx:.2f}, mm_per_px={mpp:.8f}")
        else:
            self.lbl_cal.setText("Calibration: Not calibrated (need needle).")

    def update_temp_label(self):
        d0 = None
        if self.temp_dh_mm is not None and self.temp_dv_mm is not None:
            d0 = (self.temp_dh_mm ** 2 * self.temp_dv_mm) ** (1.0 / 3.0)
        dh = "-" if self.temp_dh_mm is None else f"{self.temp_dh_mm:.5f}"
        dv = "-" if self.temp_dv_mm is None else f"{self.temp_dv_mm:.5f}"
        d0s = "-" if d0 is None else f"{d0:.5f}"
        self.lbl_tmp.setText(f"Current: Dh={dh}  Dv={dv}  D0={d0s}")

    def update_temp_label_from_saved(self):
        if self.current_image_index < 0 or not self.current_images:
            self.temp_dh_mm = None
            self.temp_dv_mm = None
            self.update_temp_label()
            return
        img_path = str(self.current_images[self.current_image_index])
        row = self.results.get(img_path)
        if row:
            self.temp_dh_mm = row.dh_mm
            self.temp_dv_mm = row.dv_mm
        else:
            self.temp_dh_mm = None
            self.temp_dv_mm = None
        self.update_temp_label()

    def current_condition_rows(self) -> List[MeasurementRow]:
        if self.current_condition is None:
            return []
        cond = self.current_condition.name
        rows = [r for r in self.results.values() if r.condition_folder == cond]
        return sorted(rows, key=lambda r: r.frame_index)

    def refresh_table(self):
        if self.current_condition is None:
            self.table.setRowCount(0)
            return

        # Refresh time columns based on current FPS/T0
        cond = self.current_condition.name
        t0_idx = self.t0_index_by_cond.get(cond)
        t0_name = self.t0_name_by_cond.get(cond)
        fps = float(self.fps_spin.value())

        for r in self.results.values():
            if r.condition_folder != cond:
                continue
            r.fps = fps
            r.t0_index = t0_idx
            r.t0_image_name = t0_name
            if t0_idx is None or fps <= 0:
                r.time_s = None
                r.time_ms = None
            else:
                r.time_s = (r.frame_index - t0_idx) / fps
                r.time_ms = r.time_s * 1000.0

        rows = self.current_condition_rows()
        self.table.setRowCount(len(rows))
        for i, r in enumerate(rows):
            self.table.setItem(i, 0, QTableWidgetItem(r.concentration))
            self.table.setItem(i, 1, QTableWidgetItem(r.speed))
            self.table.setItem(i, 2, QTableWidgetItem(r.image_name))
            self.table.setItem(i, 3, QTableWidgetItem(str(r.frame_index)))
            self.table.setItem(i, 4, QTableWidgetItem("" if r.time_s is None else f"{r.time_s:.6f}"))
            self.table.setItem(i, 5, QTableWidgetItem("" if r.dh_mm is None else f"{r.dh_mm:.5f}"))
            self.table.setItem(i, 6, QTableWidgetItem("" if r.dv_mm is None else f"{r.dv_mm:.5f}"))
            self.table.setItem(i, 7, QTableWidgetItem("" if r.d0_mm is None else f"{r.d0_mm:.5f}"))
            self.table.setItem(i, 8, QTableWidgetItem("" if r.t0_index is None else f"{r.t0_index}"))

    # ---------------- Table selection utilities ----------------
    def select_all_table(self):
        self.table.selectAll()
        self.set_status("Selected all rows in table.")

    def delete_selected_table_rows(self):
        sel = self.table.selectionModel().selectedRows()
        if not sel:
            self.set_status("No rows selected.")
            return

        rows = self.current_condition_rows()
        to_delete_paths = set()
        for idx in sel:
            r = idx.row()
            if 0 <= r < len(rows):
                to_delete_paths.add(rows[r].image_path)

        for p in to_delete_paths:
            self.results.pop(p, None)

        self.update_temp_label_from_saved()
        self.refresh_table()
        self.set_status(f"Deleted {len(to_delete_paths)} selected rows.")

    def clear_current_condition_results(self):
        if self.current_condition is None:
            return
        cond = self.current_condition.name
        to_delete = [k for k, v in self.results.items() if v.condition_folder == cond]
        for k in to_delete:
            self.results.pop(k, None)
        self.update_temp_label_from_saved()
        self.refresh_table()
        self.set_status(f"Cleared all results in condition [{cond}].")

    def clear_all_results(self):
        self.results.clear()
        self.temp_dh_mm = None
        self.temp_dv_mm = None
        self.update_temp_label()
        self.refresh_table()
        self.set_status("Cleared ALL results.")

    def on_table_context_menu(self, pos):
        menu = QMenu(self)
        act_del = menu.addAction("Delete Selected")
        act_sel_all = menu.addAction("Select All")
        menu.addSeparator()
        act_clear_cond = menu.addAction("Clear This Condition")
        act_clear_all = menu.addAction("Clear All")
        action = menu.exec(self.table.viewport().mapToGlobal(pos))
        if action == act_del:
            self.delete_selected_table_rows()
        elif action == act_sel_all:
            self.select_all_table()
        elif action == act_clear_cond:
            self.clear_current_condition_results()
        elif action == act_clear_all:
            self.clear_all_results()

    def delete_current_image_result(self):
        if self.current_image_index < 0 or not self.current_images:
            return
        img_path = str(self.current_images[self.current_image_index])
        if img_path in self.results:
            del self.results[img_path]
            self.update_temp_label_from_saved()
            self.refresh_table()
            self.set_status("Deleted saved result for this image.")
        else:
            self.set_status("No saved result for this image to delete.")

    # ---------------- Export ----------------
    def export_excel(self):
        if not self.results:
            self.message("Hint", "No results to export yet.")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "droplet_results.xlsx", "Excel (*.xlsx)")
        if not save_path:
            return
        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append([
            "ConditionFolder", "Concentration", "Speed",
            "N", "FPS", "T0_Index", "T0_ImageName",
            "Dh_avg(mm)", "Dv_avg(mm)", "D0_avg(mm)"
        ])

        by_cond: Dict[str, List[MeasurementRow]] = {}
        for r in self.results.values():
            by_cond.setdefault(r.condition_folder, []).append(r)

        for cond, rows in sorted(by_cond.items(), key=lambda x: x[0]):
            rows = sorted(rows, key=lambda r: r.frame_index)
            ws = wb.create_sheet(title=clamp_sheet_name(cond))
            ws.append([
                "Concentration", "Speed",
                "FrameIndex", "ImageName", "ImagePath",
                "FPS", "T0_Index", "T0_ImageName", "Time_s", "Time_ms",
                "Needle_px", "mm_per_px",
                "Dh(mm)", "Dv(mm)", "D0(mm)"
            ])

            for r in rows:
                ws.append([
                    r.concentration,
                    r.speed,
                    int(r.frame_index),
                    r.image_name,
                    r.image_path,
                    float(r.fps),
                    "" if r.t0_index is None else int(r.t0_index),
                    "" if r.t0_image_name is None else r.t0_image_name,
                    "" if r.time_s is None else float(r.time_s),
                    "" if r.time_ms is None else float(r.time_ms),
                    "" if r.needle_px is None else float(r.needle_px),
                    "" if r.mm_per_px is None else float(r.mm_per_px),
                    "" if r.dh_mm is None else float(round(r.dh_mm, 5)),
                    "" if r.dv_mm is None else float(round(r.dv_mm, 5)),
                    "" if r.d0_mm is None else float(round(r.d0_mm, 5)),
                ])

            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

            dhs = [r.dh_mm for r in rows if r.dh_mm is not None]
            dvs = [r.dv_mm for r in rows if r.dv_mm is not None]
            d0s = [r.d0_mm for r in rows if r.d0_mm is not None]
            n = len(d0s)

            dh_avg = sum(dhs) / len(dhs) if dhs else None
            dv_avg = sum(dvs) / len(dvs) if dvs else None
            d0_avg = sum(d0s) / len(d0s) if d0s else None

            conc, spd = parse_condition(cond)
            t0_idx = self.t0_index_by_cond.get(cond)
            t0_name = self.t0_name_by_cond.get(cond)
            fps = float(self.fps_spin.value())

            ws_summary.append([
                cond, conc, spd,
                n, fps,
                "" if t0_idx is None else int(t0_idx),
                "" if t0_name is None else t0_name,
                "" if dh_avg is None else float(round(dh_avg, 5)),
                "" if dv_avg is None else float(round(dv_avg, 5)),
                "" if d0_avg is None else float(round(d0_avg, 5)),
            ])

        try:
            wb.save(save_path)
        except Exception as e:
            self.message("Error", f"Failed to save Excel:\n{e}", icon=QMessageBox.Critical)
            return

        self.message("Done", f"Exported:\n{save_path}")
        self.set_status("Export finished.")


def main():
    app = QApplication([])
    win = MainWindow()
    win.show()
    app.exec()


if __name__ == "__main__":
    main()
