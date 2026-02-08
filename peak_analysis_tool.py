# -*- coding: utf-8 -*-
"""
峰值分析工具（双滤波）- 增强版 v3.4
界面优化：文件选择功能移至右侧功能区，移除顶部工具栏
"""

import os
import sys
import math
import traceback
import numpy as np
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Tuple, List, Dict, Any

from PySide6.QtCore import Qt, QTimer, QPoint, Signal
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QRadioButton, QCheckBox,
    QGroupBox, QFileDialog, QMessageBox, QSplitter,
    QGridLayout, QScrollArea, QSizePolicy, QComboBox,
    QDoubleSpinBox, QSpinBox, QFrame
)

import matplotlib

matplotlib.use('QtAgg')
import matplotlib.pyplot as plt

# 提前设置matplotlib参数
plt.rcParams.update({
    'font.sans-serif': ['Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans'],
    'axes.unicode_minus': False,
    'font.size': 9
})

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure

# 尝试导入scipy
try:
    from scipy.interpolate import interp1d

    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False
    print("提示: SciPy未安装，插值功能将使用numpy实现")

try:
    from openpyxl import Workbook, load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("提示: openpyxl未安装，Excel导出功能不可用")

# ========== 流体参数表 ==========
FLUID_PARAMS = {
    "water": {"rho": 997.0, "sigma": 0.0707},
    "800ppm": {"rho": 998.0, "sigma": 0.0707},
    "2000ppm": {"rho": 998.0, "sigma": 0.072},
    "4000ppm": {"rho": 998.0, "sigma": 0.072},
    "6000ppm": {"rho": 998.0, "sigma": 0.072},
    "10000ppm": {"rho": 1002.0, "sigma": 0.074}
}


# ========== 数据结构定义 ==========
@dataclass
class Params:
    """基本参数"""
    SIGMA: float = 5.0
    FIR_FC: float = 2000.0
    PRE_TIME: float = 0.005
    POST_TIME: float = 0.015
    STEP: float = 0.00001
    COEFF_A: float = 0.05
    INIT_FILE: str = ""  # 默认为空，启动后手动选择文件


@dataclass
class FluidProperties:
    """流体物性参数"""
    fluid_name: str = "2000ppm"
    rho: float = 998.0
    sigma: float = 0.072


@dataclass
class PeakInfo:
    """峰值信息"""
    is_manual: bool = False
    idx: Optional[int] = None
    T0: Optional[float] = None
    V0: Optional[float] = None
    is_peak1_manual: bool = False
    is_peak2_manual: bool = False
    peak1_time_rel: Optional[float] = None
    peak2_time_rel: Optional[float] = None


@dataclass
class CalibrationData:
    """校准数据"""
    t_raw: np.ndarray
    F_raw: np.ndarray
    t_smooth: np.ndarray
    F_smooth: np.ndarray
    t_rho: np.ndarray
    t_gamma: np.ndarray
    F_star: np.ndarray


# ========== 工具函数 ==========
def gaussian_filter_numpy(signal: np.ndarray, sigma: float) -> np.ndarray:
    """高斯滤波的numpy实现"""
    if sigma <= 0:
        return signal.copy()

    kernel_size = max(int(round(6 * sigma)) + 1, 3)
    if kernel_size % 2 == 0:
        kernel_size += 1

    x = np.arange(-(kernel_size // 2), kernel_size // 2 + 1)
    kernel = np.exp(-x ** 2 / (2 * sigma ** 2))
    kernel = kernel / np.sum(kernel)

    result = np.convolve(signal, kernel, mode='same')
    return result


def fir_filter_numpy(signal: np.ndarray, x_time: np.ndarray, fc: float) -> np.ndarray:
    """FIR滤波的numpy实现"""
    if len(x_time) < 2:
        return signal.copy()

    dt = float(x_time[1] - x_time[0])
    if dt <= 0:
        return signal.copy()

    fs = 1.0 / dt
    nyquist = fs / 2.0
    cutoff = fc / nyquist

    if cutoff >= 1.0 or cutoff <= 0.0:
        return signal.copy()

    # 设计FIR滤波器
    ntaps = 101
    if ntaps % 2 == 0:
        ntaps += 1

    m = ntaps - 1
    n = np.arange(-m // 2, m // 2 + 1)
    h = np.sinc(2 * cutoff * n)
    w = np.hamming(ntaps)
    b = h * w
    b = b / np.sum(b)

    result = np.convolve(signal, b, mode='same')
    return result


def clean_file_path(path_str: str) -> str:
    """清理文件路径字符串"""
    cleaned = str(path_str).strip()
    if cleaned.startswith('"') and cleaned.endswith('"'):
        cleaned = cleaned[1:-1]
    return cleaned


def extract_params_from_filename(filename: str) -> Tuple[str, float, float]:
    """从文件名中提取参数"""
    fluid_name, D0, U0 = "Unknown", np.nan, np.nan

    try:
        basename = os.path.basename(filename)
        name, _ = os.path.splitext(basename)
        parts = name.split('_')

        if len(parts) >= 3:
            first_part = parts[0].lower()
            matched_fluid = None
            for fluid in FLUID_PARAMS:
                if fluid in first_part:
                    matched_fluid = fluid
                    break

            if matched_fluid:
                fluid_name = matched_fluid
            else:
                fluid_name = parts[0]

            D0 = float(parts[1])
            U0 = float(parts[2])
    except Exception as e:
        print(f"从文件名提取参数出错: {e}")

    return fluid_name, D0, U0


# ========== 放大窗口类 ==========
class MagnifierWindow(QWidget):
    """放大窗口类"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_widget = parent
        self.init_ui()

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("实时放大窗口")
        self.resize(600, 450)
        self.setStyleSheet("""
            QWidget {
                background-color: white;
            }
        """)

        layout = QVBoxLayout(self)

        # 创建图形
        self.fig = Figure(figsize=(6, 4.5), dpi=100)
        self.canvas = FigureCanvas(self.fig)
        self.ax = self.fig.add_subplot(111)

        layout.addWidget(self.canvas)

    def update_magnifier(self, t_data, F_data, current_pos, label_text="放大视图"):
        """更新放大窗口内容"""
        if len(t_data) == 0:
            return

        self.ax.clear()

        # 计算放大范围
        total_range = t_data[-1] - t_data[0]
        window_size = total_range / 2.5  # 放大2.5倍

        t_min = max(t_data[0], current_pos - window_size / 2)
        t_max = min(t_data[-1], current_pos + window_size / 2)

        # 提取窗口内的数据
        mask = (t_data >= t_min) & (t_data <= t_max)
        if np.sum(mask) < 10:
            t_min = max(t_data[0], current_pos - window_size)
            t_max = min(t_data[-1], current_pos + window_size)
            mask = (t_data >= t_min) & (t_data <= t_max)

        t_window = t_data[mask]
        F_window = F_data[mask]

        if len(t_window) < 2:
            return

        # 绘制曲线
        self.ax.plot(t_window, F_window, linewidth=2.0, color='purple')

        # 标记当前点
        if len(t_window) > 0:
            idx = np.argmin(np.abs(t_window - current_pos))
            nearest_t = t_window[idx]
            nearest_F = F_window[idx]

            # 绘制标记线和点
            self.ax.axvline(x=nearest_t, color='red', linestyle='--', linewidth=1.5, alpha=0.9)
            self.ax.axhline(y=nearest_F, color='green', linestyle='--', linewidth=1.5, alpha=0.7)
            self.ax.scatter([nearest_t], [nearest_F],
                            s=120, c='red', edgecolors='white', linewidth=2.0, zorder=5)

        self.ax.axhline(y=0, color='blue', linestyle='--', linewidth=1.0, alpha=0.6)
        self.ax.set_xlabel("相对时间 (s)", fontsize=11)
        self.ax.set_ylabel("碰撞力 (N)", fontsize=11)
        self.ax.set_title(f"{label_text}: {current_pos:.6f} s", fontsize=12)
        self.ax.grid(True, alpha=0.4)
        self.ax.tick_params(labelsize=10)

        self.fig.tight_layout()
        self.canvas.draw()


# ========== 主窗口 ==========
class PeakAnalysisTool(QMainWindow):
    def __init__(self):
        super().__init__()

        # 初始化参数
        self.params = Params()
        self.fluid_props = FluidProperties()

        # 全局变量
        self.x = np.array([])
        self.y = np.array([])
        self.y_smoothed_global = np.array([])
        self.filter_type = 'fir'
        self.peak_info = PeakInfo()
        self.delta_t_user = 0.0
        self.delta_F_user = 0.0
        self.U0 = 1.2
        self.D0 = 2.15
        self.calib_data = None

        # 手动选点模式
        self.manual_peak_mode = None  # None, 'T0', 'Peak1', 'Peak2'

        # 放大窗口
        self.magnifier_window = None
        self.magnifier_active = False
        self.adjusting_peak = False  # 用于标记是否正在调整峰值

        # T0微调步长
        self.t0_adjust_step = 0.00001  # 默认步长10微秒
        self.peak_adjust_step = 0.00001  # 峰值微调步长

        # 无量纲曲线图的顶部坐标轴引用
        self.ax_dimensionless_top = None

        # 对比曲线存储列表（增强版：包含完整元数据和缓存数据）
        self.compare_curves = []  # 存储格式: [{"label": str, "t_rho": array, "t_gamma": array, "F_star": array, "visible": bool, "metadata": dict, "cached_data": dict}, ...]

        # 文件夹选择相关变量
        self.data_folder = ""           # 数据文件夹路径
        self.all_files = []             # 文件夹中所有CSV文件列表
        self.file_index = {}            # 按流体类型和速度索引的文件字典
        # 结构: {fluid_type: {velocity: [file1, file2, ...]}}
        self.current_fluid_type = ""    # 当前选择的流体类型
        self.current_velocity = ""      # 当前选择的速度
        self.current_file_list = []     # 当前流体类型+速度下的文件列表
        self.current_file_idx = 0       # 当前文件在列表中的索引
        self.current_file_path = ""     # 当前加载的文件路径

        # 初始化UI
        self.init_ui()

        # 初始加载数据
        QTimer.singleShot(100, self.load_initial_data)

    def init_ui(self):
        """初始化用户界面"""
        self.setWindowTitle('峰值分析工具（双滤波）- 增强版 v3.3')
        self.setGeometry(100, 100, 1600, 1000)

        # 设置样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 9pt;
                border: 1px solid #cccccc;
                border-radius: 4px;
                margin-top: 6px;
                padding-top: 6px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px 0 4px;
            }
            QLabel {
                font-size: 8.5pt;
                min-height: 18px;
            }
            QPushButton {
                font-size: 8.5pt;
                padding: 4px 10px;
                background-color: #e8e8e8;
                border: 1px solid #aaaaaa;
                border-radius: 3px;
                min-height: 22px;
                min-width: 60px;
            }
            QPushButton:hover {
                background-color: #d8d8d8;
            }
            QPushButton:pressed {
                background-color: #c8c8c8;
            }
            QLineEdit {
                font-size: 8.5pt;
                padding: 3px 5px;
                border: 1px solid #aaaaaa;
                border-radius: 3px;
                min-height: 22px;
                max-width: 100px;
            }
            QRadioButton {
                font-size: 8.5pt;
                spacing: 4px;
                min-height: 18px;
            }
            QComboBox {
                font-size: 8.5pt;
                padding: 3px 5px;
                border: 1px solid #aaaaaa;
                border-radius: 3px;
                min-height: 22px;
                max-width: 120px;
            }
            QDoubleSpinBox {
                font-size: 8.5pt;
                padding: 3px 5px;
                border: 1px solid #aaaaaa;
                border-radius: 3px;
                min-height: 22px;
                max-width: 120px;
            }
        """)

        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(6)

        # 分割器：左侧图表，右侧控制
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter, 1)

        # 左侧：图表区域
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        # 创建图表画布
        self.fig = Figure(figsize=(12, 10), dpi=100)
        self.canvas = FigureCanvas(self.fig)
        self.toolbar = NavigationToolbar(self.canvas, self)

        # 创建子图布局 - 使用GridSpec来调整布局
        gs = self.fig.add_gridspec(3, 2, height_ratios=[1, 1, 1.5])  # 第三行高度是前两行的1.5倍

        # 创建子图
        self.ax_raw = self.fig.add_subplot(gs[0, 0])
        self.ax_smooth = self.fig.add_subplot(gs[0, 1])
        self.ax_calib = self.fig.add_subplot(gs[1, 0])
        self.ax_calib_smooth = self.fig.add_subplot(gs[1, 1])

        # 无量纲曲线图占据第三行左列
        self.ax_dimensionless = self.fig.add_subplot(gs[2, 0])
        # 对比曲线图占据第三行右列
        self.ax_compare = self.fig.add_subplot(gs[2, 1])

        # 设置子图标题和标签
        self.configure_axes(self.ax_raw, "原始数据曲线", "时间 (s)", "电压 (V)")
        self.configure_axes(self.ax_smooth, "FIR低通滤波曲线 - 点击快速定位峰值", "时间 (s)", "电压 (V)")
        self.configure_axes(self.ax_calib, "校准曲线（原始）", "相对时间 (s)", "碰撞力 (N)")
        self.configure_axes(self.ax_calib_smooth, "校准曲线（滤波） - 点击选择点", "相对时间 (s)", "碰撞力 (N)")

        # 配置合并的无量纲图
        self.ax_dimensionless.set_title("无量纲曲线", fontsize=10, pad=12)
        self.ax_dimensionless.set_xlabel("t/τρ", fontsize=9, labelpad=8)
        self.ax_dimensionless.set_ylabel("F*", fontsize=9, labelpad=8)
        self.ax_dimensionless.grid(True, alpha=0.3)
        self.ax_dimensionless.tick_params(labelsize=8)

        # 配置对比曲线图
        self.ax_compare.set_title("对比曲线", fontsize=10, pad=12)
        self.ax_compare.set_xlabel("t/τρ", fontsize=9, labelpad=8)
        self.ax_compare.set_ylabel("F*", fontsize=9, labelpad=8)
        self.ax_compare.grid(True, alpha=0.3)
        self.ax_compare.tick_params(labelsize=8)

        # 调整布局
        self.fig.tight_layout()

        left_layout.addWidget(self.toolbar)
        left_layout.addWidget(self.canvas)

        # 右侧：控制区域 - 优化版
        right_widget = QWidget()

        # 创建滚动区域
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setWidget(right_widget)
        right_scroll.setMinimumWidth(380)
        right_scroll.setMaximumWidth(450)

        # 关键设置：禁用水平滚动条，只允许垂直滚动
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # 设置滚动区域内部部件的大小策略
        right_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 右侧布局 - 使用固定宽度并填满
        right_layout = QVBoxLayout(right_widget)
        right_layout.setSpacing(6)
        right_layout.setContentsMargins(8, 8, 8, 8)  # 减小边距

        # 设置右侧部件的尺寸策略
        right_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # ========== 文件选择区域（紧凑布局） ==========
        file_select_group = QGroupBox("文件选择")
        file_select_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        file_select_layout = QGridLayout()
        file_select_layout.setHorizontalSpacing(6)
        file_select_layout.setVerticalSpacing(4)
        file_select_layout.setContentsMargins(8, 8, 8, 8)

        # 第一行：选择文件夹按钮 + 路径显示
        self.browse_folder_btn = QPushButton("选择文件夹")
        self.browse_folder_btn.clicked.connect(self.browse_folder)
        self.browse_folder_btn.setMaximumWidth(80)
        self.folder_edit = QLineEdit()
        self.folder_edit.setReadOnly(True)
        self.folder_edit.setPlaceholderText("请选择数据文件夹...")
        file_select_layout.addWidget(self.browse_folder_btn, 0, 0)
        file_select_layout.addWidget(self.folder_edit, 0, 1, 1, 3)

        # 第二行：流体类型 + 速度（同一行）
        self.fluid_select_label = QLabel("流体:")
        self.fluid_select_combo = QComboBox()
        self.fluid_select_combo.currentTextChanged.connect(self.on_fluid_type_changed)
        self.fluid_select_combo.setEnabled(False)
        self.velocity_select_label = QLabel("速度:")
        self.velocity_select_combo = QComboBox()
        self.velocity_select_combo.currentTextChanged.connect(self.on_velocity_changed)
        self.velocity_select_combo.setEnabled(False)
        file_select_layout.addWidget(self.fluid_select_label, 1, 0)
        file_select_layout.addWidget(self.fluid_select_combo, 1, 1)
        file_select_layout.addWidget(self.velocity_select_label, 1, 2)
        file_select_layout.addWidget(self.velocity_select_combo, 1, 3)

        # 第三行：文件导航
        self.prev_btn = QPushButton("◀ 上一个")
        self.prev_btn.clicked.connect(self.load_prev_file)
        self.prev_btn.setEnabled(False)
        self.next_btn = QPushButton("下一个 ▶")
        self.next_btn.clicked.connect(self.load_next_file)
        self.next_btn.setEnabled(False)
        self.file_info_label = QLabel("")
        self.file_info_label.setStyleSheet("color: #0066cc; font-weight: bold;")
        file_select_layout.addWidget(self.prev_btn, 2, 0, 1, 2)
        file_select_layout.addWidget(self.next_btn, 2, 2, 1, 2)
        file_select_layout.addWidget(self.file_info_label, 3, 0, 1, 4)

        file_select_group.setLayout(file_select_layout)
        right_layout.addWidget(file_select_group)

        # ========== 滤波设置区域 ==========
        filter_group = QGroupBox("滤波设置")
        filter_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        filter_layout = QGridLayout()
        filter_layout.setHorizontalSpacing(8)
        filter_layout.setVerticalSpacing(6)
        filter_layout.setContentsMargins(8, 10, 8, 10)

        self.fir_rb = QRadioButton("FIR低通滤波 (默认)")
        self.gaussian_rb = QRadioButton("高斯滤波 (可选)")
        self.fir_rb.setChecked(True)
        self.gaussian_rb.setChecked(False)
        self.fir_rb.clicked.connect(lambda: self.switch_filter_type('fir'))
        self.gaussian_rb.clicked.connect(lambda: self.switch_filter_type('gaussian'))

        self.fir_fc_label = QLabel("截止频率(Hz):")
        self.fir_fc_edit = QLineEdit(str(self.params.FIR_FC))
        self.fir_fc_edit.setMaximumWidth(80)

        self.sigma_label = QLabel("Sigma:")
        self.sigma_edit = QLineEdit(str(self.params.SIGMA))
        self.sigma_edit.setMaximumWidth(80)
        self.sigma_edit.setEnabled(False)

        self.apply_filter_btn = QPushButton("应用滤波")
        self.apply_filter_btn.clicked.connect(self.reapply_filter)

        filter_layout.addWidget(self.fir_rb, 0, 0, 1, 2)
        filter_layout.addWidget(self.gaussian_rb, 0, 2, 1, 2)
        filter_layout.addWidget(self.fir_fc_label, 1, 0)
        filter_layout.addWidget(self.fir_fc_edit, 1, 1)
        filter_layout.addWidget(self.sigma_label, 1, 2)
        filter_layout.addWidget(self.sigma_edit, 1, 3)
        filter_layout.addWidget(self.apply_filter_btn, 2, 0, 1, 4)

        filter_group.setLayout(filter_layout)
        right_layout.addWidget(filter_group)

        # ========== 流体参数区域 ==========
        fluid_group = QGroupBox("流体参数")
        fluid_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        fluid_layout = QGridLayout()
        fluid_layout.setHorizontalSpacing(8)
        fluid_layout.setVerticalSpacing(6)
        fluid_layout.setContentsMargins(8, 10, 8, 10)

        self.fluid_name_label = QLabel("流体名称:")
        self.fluid_name_combo = QComboBox()
        self.fluid_name_combo.addItems(list(FLUID_PARAMS.keys()))
        self.fluid_name_combo.currentTextChanged.connect(self.update_fluid_params)
        self.fluid_name_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.rho_label = QLabel("密度 (kg/m³):")
        self.rho_edit = QLineEdit(str(self.fluid_props.rho))
        self.rho_edit.setMaximumWidth(80)
        self.rho_edit.textChanged.connect(self.schedule_update)

        self.sigma_label_fluid = QLabel("表面张力 (N/m):")
        self.sigma_edit_fluid = QLineEdit(str(self.fluid_props.sigma))
        self.sigma_edit_fluid.setMaximumWidth(80)
        self.sigma_edit_fluid.textChanged.connect(self.schedule_update)

        self.u0_label = QLabel("U0 (m/s):")
        self.u0_edit = QLineEdit(str(self.U0))
        self.u0_edit.setMaximumWidth(80)
        self.u0_edit.textChanged.connect(self.schedule_update)

        self.d0_label = QLabel("D0 (mm):")
        self.d0_edit = QLineEdit(str(self.D0))
        self.d0_edit.setMaximumWidth(80)
        self.d0_edit.textChanged.connect(self.schedule_update)

        fluid_layout.addWidget(self.fluid_name_label, 0, 0)
        fluid_layout.addWidget(self.fluid_name_combo, 0, 1, 1, 3)
        fluid_layout.addWidget(self.rho_label, 1, 0)
        fluid_layout.addWidget(self.rho_edit, 1, 1)
        fluid_layout.addWidget(self.sigma_label_fluid, 1, 2)
        fluid_layout.addWidget(self.sigma_edit_fluid, 1, 3)
        fluid_layout.addWidget(self.u0_label, 2, 0)
        fluid_layout.addWidget(self.u0_edit, 2, 1)
        fluid_layout.addWidget(self.d0_label, 2, 2)
        fluid_layout.addWidget(self.d0_edit, 2, 3)

        fluid_group.setLayout(fluid_layout)
        right_layout.addWidget(fluid_group)

        # ========== 快速定位区域 ==========
        quick_locate_group = QGroupBox("快速定位峰值")
        quick_locate_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        quick_locate_layout = QVBoxLayout()
        quick_locate_layout.setSpacing(6)
        quick_locate_layout.setContentsMargins(8, 10, 8, 10)

        quick_locate_instruction = QLabel(
            "在左侧'FIR低通滤波曲线'图中点击任意位置，程序将自动定位该位置附近的最高峰值"
        )
        quick_locate_instruction.setStyleSheet("color: #666666; font-size: 8pt; font-style: italic;")
        quick_locate_instruction.setWordWrap(True)
        quick_locate_layout.addWidget(quick_locate_instruction)

        quick_locate_frame = QWidget()
        quick_locate_buttons_layout = QHBoxLayout(quick_locate_frame)
        quick_locate_buttons_layout.setContentsMargins(0, 0, 0, 0)
        quick_locate_buttons_layout.setSpacing(6)

        self.quick_locate_label = QLabel("快速定位:")
        self.quick_locate_info = QLabel("未定位")
        self.quick_locate_info.setStyleSheet("color: #0066cc;")
        self.quick_locate_info.setWordWrap(True)

        quick_locate_buttons_layout.addWidget(self.quick_locate_label)
        quick_locate_buttons_layout.addWidget(self.quick_locate_info, 1)

        quick_locate_layout.addWidget(quick_locate_frame)
        quick_locate_group.setLayout(quick_locate_layout)
        right_layout.addWidget(quick_locate_group)

        # ========== T0精确校准区域 ==========
        t0_group = QGroupBox("T0精确校准")
        t0_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        t0_layout = QVBoxLayout()
        t0_layout.setSpacing(6)
        t0_layout.setContentsMargins(8, 10, 8, 10)

        # 当前T0显示
        t0_display_frame = QWidget()
        t0_display_layout = QHBoxLayout(t0_display_frame)
        t0_display_layout.setContentsMargins(0, 0, 0, 0)
        t0_display_layout.setSpacing(6)

        self.t0_label = QLabel("当前T0:")
        self.t0_value_label = QLabel("未设置")
        self.t0_value_label.setStyleSheet("font-weight: bold; color: #d00000;")
        self.t0_value_label.setWordWrap(True)
        t0_display_layout.addWidget(self.t0_label)
        t0_display_layout.addWidget(self.t0_value_label, 1)
        t0_display_layout.addStretch()

        # 微调步长设置
        step_frame = QWidget()
        step_layout = QHBoxLayout(step_frame)
        step_layout.setContentsMargins(0, 0, 0, 0)
        step_layout.setSpacing(6)

        self.step_label = QLabel("微调步长(s):")
        self.step_spinbox = QDoubleSpinBox()
        self.step_spinbox.setDecimals(7)
        self.step_spinbox.setMinimum(0.0000001)
        self.step_spinbox.setMaximum(0.01)
        self.step_spinbox.setSingleStep(0.000001)
        self.step_spinbox.setValue(self.t0_adjust_step)
        self.step_spinbox.setSuffix(" s")
        self.step_spinbox.valueChanged.connect(self.update_t0_adjust_step)
        self.step_spinbox.setMaximumWidth(120)

        step_layout.addWidget(self.step_label)
        step_layout.addWidget(self.step_spinbox, 1)

        # 微调按钮
        adjust_frame = QWidget()
        adjust_layout = QHBoxLayout(adjust_frame)
        adjust_layout.setContentsMargins(0, 0, 0, 0)
        adjust_layout.setSpacing(6)

        self.adjust_left_btn = QPushButton("← 左移")
        self.adjust_left_btn.clicked.connect(lambda: self.adjust_t0(-1))
        self.adjust_left_btn.setEnabled(False)

        self.adjust_right_btn = QPushButton("右移 →")
        self.adjust_right_btn.clicked.connect(lambda: self.adjust_t0(1))
        self.adjust_right_btn.setEnabled(False)

        adjust_layout.addWidget(self.adjust_left_btn)
        adjust_layout.addWidget(self.adjust_right_btn)

        # 校准按钮
        t0_buttons_frame = QWidget()
        t0_buttons_layout = QHBoxLayout(t0_buttons_frame)
        t0_buttons_layout.setContentsMargins(0, 0, 0, 0)
        t0_buttons_layout.setSpacing(6)

        self.calibrate_t0_btn = QPushButton("精确校准T0")
        self.calibrate_t0_btn.clicked.connect(self.start_t0_calibration)
        self.cancel_calibrate_btn = QPushButton("取消")
        self.cancel_calibrate_btn.clicked.connect(self.cancel_calibration)
        self.cancel_calibrate_btn.setEnabled(False)

        t0_buttons_layout.addWidget(self.calibrate_t0_btn)
        t0_buttons_layout.addWidget(self.cancel_calibrate_btn)
        t0_buttons_layout.addStretch()

        t0_layout.addWidget(t0_display_frame)
        t0_layout.addWidget(step_frame)
        t0_layout.addWidget(adjust_frame)
        t0_layout.addWidget(t0_buttons_frame)

        # 校准说明
        self.calibration_instruction = QLabel(
            "提示：点击'精确校准T0'按钮后，在'校准曲线（滤波）'图中点击选择零点位置\n"
            "鼠标悬停时会显示放大窗口，便于精确选择\n"
            "或者使用微调按钮进行精细调整（微调时会显示放大窗口）"
        )
        self.calibration_instruction.setStyleSheet("color: #666666; font-size: 8pt; font-style: italic;")
        self.calibration_instruction.setWordWrap(True)
        t0_layout.addWidget(self.calibration_instruction)

        t0_group.setLayout(t0_layout)
        right_layout.addWidget(t0_group)

        # ========== 手动峰值设置区域 ==========
        manual_group = QGroupBox("峰值设置（带放大窗口）")
        manual_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        manual_layout = QVBoxLayout()
        manual_layout.setSpacing(6)
        manual_layout.setContentsMargins(8, 10, 8, 10)

        # Peak1选择
        peak1_frame = QWidget()
        peak1_layout = QHBoxLayout(peak1_frame)
        peak1_layout.setContentsMargins(0, 0, 0, 0)
        peak1_layout.setSpacing(6)

        self.peak1_label = QLabel("Peak1:")
        self.peak1_edit = QLineEdit()
        self.peak1_edit.setMaximumWidth(100)

        # Peak1微调按钮
        self.peak1_left_btn = QPushButton("←")
        self.peak1_left_btn.clicked.connect(lambda: self.adjust_peak_time_with_magnifier('Peak1', -1))
        self.peak1_left_btn.setMaximumWidth(30)
        self.peak1_left_btn.setEnabled(False)

        self.peak1_right_btn = QPushButton("→")
        self.peak1_right_btn.clicked.connect(lambda: self.adjust_peak_time_with_magnifier('Peak1', 1))
        self.peak1_right_btn.setMaximumWidth(30)
        self.peak1_right_btn.setEnabled(False)

        self.pick_peak1_btn = QPushButton("点选Peak1")
        self.pick_peak1_btn.clicked.connect(lambda: self.start_manual_peak_picking('Peak1'))
        self.auto_peak1_btn = QPushButton("自动")
        self.auto_peak1_btn.clicked.connect(lambda: self.auto_detect_peak('Peak1'))

        peak1_layout.addWidget(self.peak1_label)
        peak1_layout.addWidget(self.peak1_edit, 1)
        peak1_layout.addWidget(self.peak1_left_btn)
        peak1_layout.addWidget(self.peak1_right_btn)
        peak1_layout.addWidget(self.pick_peak1_btn)
        peak1_layout.addWidget(self.auto_peak1_btn)

        # Peak2选择
        peak2_frame = QWidget()
        peak2_layout = QHBoxLayout(peak2_frame)
        peak2_layout.setContentsMargins(0, 0, 0, 0)
        peak2_layout.setSpacing(6)

        self.peak2_label = QLabel("Peak2:")
        self.peak2_edit = QLineEdit()
        self.peak2_edit.setMaximumWidth(100)

        # Peak2微调按钮
        self.peak2_left_btn = QPushButton("←")
        self.peak2_left_btn.clicked.connect(lambda: self.adjust_peak_time_with_magnifier('Peak2', -1))
        self.peak2_left_btn.setMaximumWidth(30)
        self.peak2_left_btn.setEnabled(False)

        self.peak2_right_btn = QPushButton("→")
        self.peak2_right_btn.clicked.connect(lambda: self.adjust_peak_time_with_magnifier('Peak2', 1))
        self.peak2_right_btn.setMaximumWidth(30)
        self.peak2_right_btn.setEnabled(False)

        self.pick_peak2_btn = QPushButton("点选Peak2")
        self.pick_peak2_btn.clicked.connect(lambda: self.start_manual_peak_picking('Peak2'))
        self.auto_peak2_btn = QPushButton("自动")
        self.auto_peak2_btn.clicked.connect(lambda: self.auto_detect_peak('Peak2'))

        peak2_layout.addWidget(self.peak2_label)
        peak2_layout.addWidget(self.peak2_edit, 1)
        peak2_layout.addWidget(self.peak2_left_btn)
        peak2_layout.addWidget(self.peak2_right_btn)
        peak2_layout.addWidget(self.pick_peak2_btn)
        peak2_layout.addWidget(self.auto_peak2_btn)

        # 峰值微调步长
        peak_step_frame = QWidget()
        peak_step_layout = QHBoxLayout(peak_step_frame)
        peak_step_layout.setContentsMargins(0, 0, 0, 0)
        peak_step_layout.setSpacing(6)

        self.peak_step_label = QLabel("峰值微调步长(s):")
        self.peak_step_spinbox = QDoubleSpinBox()
        self.peak_step_spinbox.setDecimals(7)
        self.peak_step_spinbox.setMinimum(0.0000001)
        self.peak_step_spinbox.setMaximum(0.01)
        self.peak_step_spinbox.setSingleStep(0.000001)
        self.peak_step_spinbox.setValue(self.peak_adjust_step)
        self.peak_step_spinbox.setSuffix(" s")
        self.peak_step_spinbox.valueChanged.connect(self.update_peak_adjust_step)
        self.peak_step_spinbox.setMaximumWidth(120)

        peak_step_layout.addWidget(self.peak_step_label)
        peak_step_layout.addWidget(self.peak_step_spinbox, 1)

        # 应用按钮
        apply_peaks_frame = QWidget()
        apply_peaks_layout = QHBoxLayout(apply_peaks_frame)
        apply_peaks_layout.setContentsMargins(0, 0, 0, 0)
        apply_peaks_layout.setSpacing(6)

        self.apply_peaks_btn = QPushButton("应用峰值")
        self.apply_peaks_btn.clicked.connect(self.apply_manual_peaks)
        self.auto_peaks_btn = QPushButton("自动提取")
        self.auto_peaks_btn.clicked.connect(self.extract_peaks_callback)

        apply_peaks_layout.addWidget(self.apply_peaks_btn)
        apply_peaks_layout.addWidget(self.auto_peaks_btn)

        manual_layout.addWidget(peak1_frame)
        manual_layout.addWidget(peak2_frame)
        manual_layout.addWidget(peak_step_frame)
        manual_layout.addWidget(apply_peaks_frame)

        # 峰值选择说明
        peak_instruction = QLabel(
            "提示：点击'点选Peak1'或'点选Peak2'按钮后，在'校准曲线（滤波）'图中点击选择峰值位置\n"
            "鼠标悬停时会显示放大窗口，便于精确选择\n"
            "微调按钮在调整时会自动显示放大窗口"
        )
        peak_instruction.setStyleSheet("color: #666666; font-size: 8pt; font-style: italic;")
        peak_instruction.setWordWrap(True)
        manual_layout.addWidget(peak_instruction)

        # 导出峰值按钮
        self.export_peaks_btn = QPushButton("导出峰值")
        self.export_peaks_btn.clicked.connect(self.export_peaks_callback)
        manual_layout.addWidget(self.export_peaks_btn)

        manual_group.setLayout(manual_layout)
        right_layout.addWidget(manual_group)

        # ========== 结果展示区域 ==========
        results_group = QGroupBox("结果")
        results_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        results_layout = QGridLayout()
        results_layout.setHorizontalSpacing(8)
        results_layout.setVerticalSpacing(6)
        results_layout.setContentsMargins(8, 10, 8, 10)

        # 表头
        results_layout.addWidget(QLabel("参数"), 0, 0)
        results_layout.addWidget(QLabel("t/τρ"), 0, 1)
        results_layout.addWidget(QLabel("t/τγ"), 0, 2)
        results_layout.addWidget(QLabel("F*"), 0, 3)

        # Peak1
        results_layout.addWidget(QLabel("Peak1"), 1, 0)
        self.peak1_trho_edit = QLineEdit()
        self.peak1_trho_edit.setReadOnly(True)
        self.peak1_trho_edit.setMaximumWidth(100)
        results_layout.addWidget(self.peak1_trho_edit, 1, 1)

        self.peak1_tgamma_edit = QLineEdit()
        self.peak1_tgamma_edit.setReadOnly(True)
        self.peak1_tgamma_edit.setMaximumWidth(100)
        results_layout.addWidget(self.peak1_tgamma_edit, 1, 2)

        self.peak1_fstar_edit = QLineEdit()
        self.peak1_fstar_edit.setReadOnly(True)
        self.peak1_fstar_edit.setMaximumWidth(100)
        results_layout.addWidget(self.peak1_fstar_edit, 1, 3)

        # Peak2
        results_layout.addWidget(QLabel("Peak2"), 2, 0)
        self.peak2_trho_edit = QLineEdit()
        self.peak2_trho_edit.setReadOnly(True)
        self.peak2_trho_edit.setMaximumWidth(100)
        results_layout.addWidget(self.peak2_trho_edit, 2, 1)

        self.peak2_tgamma_edit = QLineEdit()
        self.peak2_tgamma_edit.setReadOnly(True)
        self.peak2_tgamma_edit.setMaximumWidth(100)
        results_layout.addWidget(self.peak2_tgamma_edit, 2, 2)

        self.peak2_fstar_edit = QLineEdit()
        self.peak2_fstar_edit.setReadOnly(True)
        self.peak2_fstar_edit.setMaximumWidth(100)
        results_layout.addWidget(self.peak2_fstar_edit, 2, 3)

        # 韦伯数
        results_layout.addWidget(QLabel("We"), 3, 0)
        self.weber_edit = QLineEdit()
        self.weber_edit.setReadOnly(True)
        self.weber_edit.setMaximumWidth(100)
        results_layout.addWidget(self.weber_edit, 3, 1, 1, 3)

        results_group.setLayout(results_layout)
        right_layout.addWidget(results_group)

        # ========== 曲线对比区域 ==========
        compare_group = QGroupBox("曲线对比")
        compare_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        compare_layout = QVBoxLayout()
        compare_layout.setSpacing(6)
        compare_layout.setContentsMargins(8, 10, 8, 10)

        # 按钮行
        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.add_compare_btn = QPushButton("添加")
        self.add_compare_btn.clicked.connect(self.add_to_compare)
        self.export_compare_btn = QPushButton("导出对比数据（8个Sheet）")
        self.export_compare_btn.clicked.connect(self.export_compare_data)
        self.clear_compare_btn = QPushButton("清空")
        self.clear_compare_btn.clicked.connect(self.clear_compare)
        btn_row.addWidget(self.add_compare_btn)
        btn_row.addWidget(self.export_compare_btn)
        btn_row.addWidget(self.clear_compare_btn)
        compare_layout.addLayout(btn_row)

        # 曲线列表区域（使用QScrollArea）
        self.curve_list_scroll = QScrollArea()
        self.curve_list_scroll.setWidgetResizable(True)
        self.curve_list_scroll.setMaximumHeight(120)
        self.curve_list_scroll.setMinimumHeight(60)
        self.curve_list_scroll.setStyleSheet("QScrollArea { border: 1px solid #ccc; }")

        self.curve_list_widget = QWidget()
        self.curve_list_layout = QVBoxLayout(self.curve_list_widget)
        self.curve_list_layout.setSpacing(2)
        self.curve_list_layout.setContentsMargins(4, 4, 4, 4)
        self.curve_list_layout.addStretch()

        self.curve_list_scroll.setWidget(self.curve_list_widget)
        compare_layout.addWidget(self.curve_list_scroll)

        self.compare_count_label = QLabel("已添加: 0 条曲线")
        self.compare_count_label.setStyleSheet("color: #666666; font-size: 8pt;")
        compare_layout.addWidget(self.compare_count_label)

        compare_group.setLayout(compare_layout)
        right_layout.addWidget(compare_group)

        # 添加弹性空间，确保内容在顶部对齐
        right_layout.addStretch(1)

        splitter.addWidget(left_widget)
        splitter.addWidget(right_scroll)
        splitter.setSizes([1000, 400])

        # 连接鼠标事件
        self.canvas.mpl_connect('button_press_event', self.on_canvas_click)
        self.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)

    # ========== UI辅助函数 ==========
    def configure_axes(self, ax, title, xlabel, ylabel):
        """配置坐标轴格式"""
        ax.set_title(title, fontsize=9, pad=6)
        ax.set_xlabel(xlabel, fontsize=8)
        ax.set_ylabel(ylabel, fontsize=8)
        ax.grid(True, alpha=0.3)
        ax.tick_params(labelsize=8)

    def browse_folder(self):
        """选择数据文件夹"""
        try:
            folder_path = QFileDialog.getExistingDirectory(
                self, "选择数据文件夹",
                self.data_folder if self.data_folder else ""
            )
            if folder_path:
                self.data_folder = folder_path
                self.folder_edit.setText(folder_path)
                self.scan_folder(folder_path)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"选择文件夹失败:\n{str(e)}")

    def scan_folder(self, folder_path):
        """扫描文件夹并构建文件索引"""
        try:
            # 清空现有数据
            self.all_files = []
            self.file_index = {}
            self.current_file_list = []
            self.current_file_idx = 0

            # 扫描CSV文件
            csv_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.csv')]
            if not csv_files:
                QMessageBox.warning(self, "提示", "所选文件夹中没有CSV文件")
                self.fluid_select_combo.clear()
                self.velocity_select_combo.clear()
                self.fluid_select_combo.setEnabled(False)
                self.velocity_select_combo.setEnabled(False)
                return

            self.all_files = csv_files

            # 解析文件名并构建索引
            for filename in csv_files:
                fluid_name, D0, U0 = extract_params_from_filename(filename)
                if fluid_name == "Unknown" or np.isnan(U0):
                    continue

                # 使用速度作为字符串键（保留原始精度）
                velocity_str = f"{U0:.2f}"

                if fluid_name not in self.file_index:
                    self.file_index[fluid_name] = {}

                if velocity_str not in self.file_index[fluid_name]:
                    self.file_index[fluid_name][velocity_str] = []

                self.file_index[fluid_name][velocity_str].append(filename)

            # 对每个速度下的文件列表排序
            for fluid in self.file_index:
                for velocity in self.file_index[fluid]:
                    self.file_index[fluid][velocity].sort()

            # 更新流体类型下拉框
            self.fluid_select_combo.blockSignals(True)
            self.fluid_select_combo.clear()
            fluid_types = sorted(self.file_index.keys())
            self.fluid_select_combo.addItems(fluid_types)
            self.fluid_select_combo.blockSignals(False)
            self.fluid_select_combo.setEnabled(True)

            # 清空速度下拉框
            self.velocity_select_combo.blockSignals(True)
            self.velocity_select_combo.clear()
            self.velocity_select_combo.blockSignals(False)
            self.velocity_select_combo.setEnabled(False)

            # 禁用导航按钮
            self.prev_btn.setEnabled(False)
            self.next_btn.setEnabled(False)
            self.file_info_label.setText("")

            self.statusBar().showMessage(f"已扫描文件夹，找到 {len(csv_files)} 个CSV文件，{len(fluid_types)} 种流体类型")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"扫描文件夹失败:\n{str(e)}")

    def on_fluid_type_changed(self, fluid_type):
        """流体类型选择变化处理"""
        if not fluid_type or fluid_type not in self.file_index:
            return

        try:
            self.current_fluid_type = fluid_type

            # 更新速度下拉框
            self.velocity_select_combo.blockSignals(True)
            self.velocity_select_combo.clear()

            velocities = sorted(self.file_index[fluid_type].keys(), key=lambda x: float(x))
            self.velocity_select_combo.addItems(velocities)

            self.velocity_select_combo.blockSignals(False)
            self.velocity_select_combo.setEnabled(True)

            # 清空当前文件列表
            self.current_file_list = []
            self.current_file_idx = 0
            self.current_velocity = ""

            # 禁用导航按钮
            self.prev_btn.setEnabled(False)
            self.next_btn.setEnabled(False)
            self.file_info_label.setText("")

            # 自动选择并加载第一个速度
            if len(velocities) > 0:
                # 设置下拉框选中第一个速度
                self.velocity_select_combo.setCurrentIndex(0)
                # 直接调用速度变化处理函数，确保数据被加载
                self.on_velocity_changed(velocities[0])
                self.statusBar().showMessage(f"已选择流体类型: {fluid_type}，自动加载速度: {velocities[0]}")
            else:
                self.statusBar().showMessage(f"已选择流体类型: {fluid_type}，但没有可用的速度数据")

        except Exception as e:
            print(f"流体类型变化处理出错: {e}")

    def on_velocity_changed(self, velocity):
        """速度选择变化处理"""
        if not velocity or not self.current_fluid_type:
            return

        try:
            self.current_velocity = velocity

            # 获取当前流体类型和速度下的文件列表
            if self.current_fluid_type in self.file_index and velocity in self.file_index[self.current_fluid_type]:
                self.current_file_list = self.file_index[self.current_fluid_type][velocity]
                self.current_file_idx = 0

                # 更新导航按钮状态
                self.update_nav_buttons()

                # 加载第一个文件
                if self.current_file_list:
                    file_path = os.path.join(self.data_folder, self.current_file_list[0])
                    self.load_data(file_path)

        except Exception as e:
            print(f"速度变化处理出错: {e}")

    def update_nav_buttons(self):
        """更新导航按钮状态"""
        file_count = len(self.current_file_list)

        if file_count == 0:
            self.prev_btn.setEnabled(False)
            self.next_btn.setEnabled(False)
            self.file_info_label.setText("")
        else:
            self.prev_btn.setEnabled(self.current_file_idx > 0)
            self.next_btn.setEnabled(self.current_file_idx < file_count - 1)
            self.file_info_label.setText(f"{self.current_file_idx + 1}/{file_count}")

    def update_fluid_params(self, fluid_name):
        """更新流体参数"""
        try:
            if fluid_name in FLUID_PARAMS:
                params = FLUID_PARAMS[fluid_name]
                self.rho_edit.setText(str(params["rho"]))
                self.sigma_edit_fluid.setText(str(params["sigma"]))

                self.fluid_props.fluid_name = fluid_name
                self.fluid_props.rho = params["rho"]
                self.fluid_props.sigma = params["sigma"]

                self.update_all_plots()
        except Exception as e:
            print(f"更新流体参数出错: {e}")

    def load_data_wrapper(self):
        """包装加载数据函数（保留兼容性）"""
        pass

    def schedule_update(self):
        """延迟更新图表"""
        QTimer.singleShot(300, self.update_all_plots)

    def update_t0_adjust_step(self, value):
        """更新T0微调步长"""
        self.t0_adjust_step = value

    def update_peak_adjust_step(self, value):
        """更新峰值微调步长"""
        self.peak_adjust_step = value

    # ========== 核心功能函数 ==========
    def load_initial_data(self):
        """加载初始数据"""
        try:
            if os.path.exists(self.params.INIT_FILE):
                QTimer.singleShot(200, lambda: self.load_data(self.params.INIT_FILE))
        except Exception as e:
            print(f"加载初始数据出错: {e}")

    def load_data(self, file_path):
        """加载数据文件"""
        try:
            if not file_path or not os.path.exists(file_path):
                return

            file_path = clean_file_path(file_path)
            self.current_file_path = file_path

            # 从文件名提取参数
            fluid_name, D0_extracted, U0_extracted = extract_params_from_filename(file_path)
            if not np.isnan(U0_extracted) and not np.isnan(D0_extracted):
                self.U0 = float(U0_extracted)
                self.D0 = float(D0_extracted)
                self.u0_edit.setText(f"{self.U0:.3f}")
                self.d0_edit.setText(f"{self.D0:.3f}")

                if fluid_name in FLUID_PARAMS:
                    self.fluid_name_combo.setCurrentText(fluid_name)

            # 读取CSV文件
            data = np.loadtxt(file_path, delimiter=',', skiprows=9, encoding='utf-8-sig')

            if data.ndim == 1:
                raise ValueError("数据列数不足")

            self.x = data[:, 0].astype(float)
            self.y = data[:, 1].astype(float)

            if len(self.x) < 100:
                raise ValueError("数据点数太少")

            # 应用滤波
            sigma_val = float(self.sigma_edit.text() or self.params.SIGMA)
            fir_fc_val = float(self.fir_fc_edit.text() or self.params.FIR_FC)

            self.y_smoothed_global = fir_filter_numpy(self.y, self.x, fir_fc_val)

            if self.filter_type == 'gaussian' or self.gaussian_rb.isChecked():
                self.y_smoothed_global = gaussian_filter_numpy(self.y_smoothed_global, sigma_val)

            # 自动识别峰值
            peak_idx, T0, V0 = self.auto_find_peak(self.x, self.y_smoothed_global)

            self.peak_info = PeakInfo(
                is_manual=False,
                idx=peak_idx,
                T0=T0,
                V0=V0
            )

            # 重置
            self.delta_t_user = 0.0
            self.delta_F_user = 0.0
            self.peak1_edit.clear()
            self.peak2_edit.clear()
            self.clear_peak_displays()

            # 更新T0显示
            self.t0_value_label.setText(f"{T0:.6f} s")
            self.quick_locate_info.setText(f"已自动定位: {T0:.6f} s")

            # 启用微调按钮
            self.adjust_left_btn.setEnabled(True)
            self.adjust_right_btn.setEnabled(True)

            # 退出校准模式
            self.manual_peak_mode = None
            self.magnifier_active = False
            self.adjusting_peak = False
            self.calibrate_t0_btn.setEnabled(True)
            self.cancel_calibrate_btn.setEnabled(False)
            self.calibration_instruction.setText(
                "提示：点击'精确校准T0'按钮后，在'校准曲线（滤波）'图中点击选择零点位置\n"
                "鼠标悬停时会显示放大窗口，便于精确选择\n"
                "或者使用微调按钮进行精细调整（微调时会显示放大窗口）"
            )

            # 关闭放大窗口
            self.close_magnifier()

            # 更新导航按钮状态
            self.update_nav_buttons()

            # 更新图表
            QTimer.singleShot(100, self.update_all_plots)

            self.statusBar().showMessage(f'已加载: {os.path.basename(file_path)}')

        except Exception as e:
            QMessageBox.critical(self, "加载失败", f"错误: {str(e)}")

    def load_prev_file(self):
        """加载上一个文件"""
        try:
            if not self.current_file_list or self.current_file_idx <= 0:
                self.statusBar().showMessage("已经是第一个文件")
                return

            self.current_file_idx -= 1
            file_path = os.path.join(self.data_folder, self.current_file_list[self.current_file_idx])
            self.load_data(file_path)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载上一个文件失败: {str(e)}")

    def load_next_file(self):
        """加载下一个文件"""
        try:
            if not self.current_file_list or self.current_file_idx >= len(self.current_file_list) - 1:
                self.statusBar().showMessage("已经是最后一个文件")
                return

            self.current_file_idx += 1
            file_path = os.path.join(self.data_folder, self.current_file_list[self.current_file_idx])
            self.load_data(file_path)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载下一个文件失败: {str(e)}")

    def auto_find_peak(self, x, y_signal):
        """自动寻找峰值"""
        data_length = len(y_signal)
        start_idx = int(np.ceil(0.05 * data_length))
        end_idx = int(np.floor(0.95 * data_length))

        if start_idx >= end_idx:
            start_idx = 0
            end_idx = data_length

        segment = y_signal[start_idx:end_idx]
        if len(segment) == 0:
            return 0, x[0], y_signal[0]

        local_peak_idx = int(np.argmax(segment))
        peak_idx = start_idx + local_peak_idx

        T0 = float(x[peak_idx])
        V0 = float(y_signal[peak_idx])

        return peak_idx, T0, V0

    def find_local_peak(self, x, y_signal, center_time, search_window=0.001):
        """在指定位置附近寻找局部最高峰值"""
        # 计算搜索范围
        start_time = center_time - search_window
        end_time = center_time + search_window

        # 找到时间范围内的数据点
        mask = (x >= start_time) & (x <= end_time)

        if np.sum(mask) == 0:
            # 如果没有找到数据点，返回中心点
            idx = int(np.argmin(np.abs(x - center_time)))
            return idx, x[idx], y_signal[idx]

        # 找到范围内的最高点
        x_range = x[mask]
        y_range = y_signal[mask]

        local_peak_idx = int(np.argmax(y_range))

        # 转换为全局索引
        global_indices = np.where(mask)[0]
        global_idx = global_indices[local_peak_idx]

        T0 = float(x_range[local_peak_idx])
        V0 = float(y_range[local_peak_idx])

        return global_idx, T0, V0

    def switch_filter_type(self, new_type):
        """切换滤波类型"""
        if self.filter_type == new_type:
            return

        self.filter_type = new_type

        if new_type == 'fir':
            self.fir_rb.setChecked(True)
            self.gaussian_rb.setChecked(False)
            self.fir_fc_edit.setEnabled(True)
            self.sigma_edit.setEnabled(False)
        else:
            self.fir_rb.setChecked(False)
            self.gaussian_rb.setChecked(True)
            self.fir_fc_edit.setEnabled(False)
            self.sigma_edit.setEnabled(True)

        self.reapply_filter()

    def reapply_filter(self):
        """重新应用滤波"""
        if len(self.x) == 0 or len(self.y) == 0:
            return

        # 获取参数
        try:
            sigma_val = float(self.sigma_edit.text() or self.params.SIGMA)
            fir_fc_val = float(self.fir_fc_edit.text() or self.params.FIR_FC)
        except ValueError:
            QMessageBox.warning(self, "警告", "滤波参数必须为数字")
            return

        # 更新参数
        self.params.SIGMA = sigma_val
        self.params.FIR_FC = fir_fc_val

        # 应用滤波
        self.y_smoothed_global = fir_filter_numpy(self.y, self.x, fir_fc_val)

        if self.filter_type == 'gaussian' or self.gaussian_rb.isChecked():
            self.y_smoothed_global = gaussian_filter_numpy(self.y_smoothed_global, sigma_val)

        # 更新峰值
        if not self.peak_info.is_manual:
            peak_idx, T0, V0 = self.auto_find_peak(self.x, self.y_smoothed_global)
            self.peak_info.idx = peak_idx
            self.peak_info.T0 = T0
            self.peak_info.V0 = V0
            self.t0_value_label.setText(f"{T0:.6f} s")

        # 更新所有图表
        self.update_all_plots()

    def update_all_plots(self):
        """更新所有图表"""
        try:
            if self.peak_info is None or self.peak_info.T0 is None:
                return

            # 获取参数
            try:
                U0_new = float(self.u0_edit.text() or self.U0)
                D0_new = float(self.d0_edit.text() or self.D0)
                rho_new = float(self.rho_edit.text() or self.fluid_props.rho)
                sigma_new = float(self.sigma_edit_fluid.text() or self.fluid_props.sigma)
            except ValueError:
                return

            # 更新参数
            self.U0 = U0_new
            self.D0 = D0_new
            self.fluid_props.rho = rho_new
            self.fluid_props.sigma = sigma_new

            # 获取校准数据
            t_final_raw, F_final_raw = self.get_calib_data(
                self.params,
                float(self.peak_info.T0),
                float(self.peak_info.V0),
                self.delta_t_user,
                self.delta_F_user,
                self.y
            )

            t_final_smooth, F_final_smooth = self.get_calib_data(
                self.params,
                float(self.peak_info.T0),
                float(self.peak_info.V0),
                self.delta_t_user,
                self.delta_F_user,
                self.y_smoothed_global
            )

            if len(t_final_raw) == 0 or len(t_final_smooth) == 0:
                return

            # 计算无量纲参数
            D0_m = self.D0 / 1000.0

            F_star_raw = F_final_raw / (self.fluid_props.rho * self.U0 ** 2 * D0_m ** 2)
            F_star_smooth = F_final_smooth / (self.fluid_props.rho * self.U0 ** 2 * D0_m ** 2)

            t_tau_rho_raw = (t_final_raw * self.U0) / D0_m
            t_tau_rho_smooth = (t_final_smooth * self.U0) / D0_m

            tau_gamma = np.sqrt(self.fluid_props.rho * D0_m ** 3 / self.fluid_props.sigma)
            t_tau_gamma_raw = t_final_raw / tau_gamma
            t_tau_gamma_smooth = t_final_smooth / tau_gamma

            # 计算韦伯数
            We = (self.fluid_props.rho * self.U0 ** 2 * D0_m) / self.fluid_props.sigma
            self.weber_edit.setText(f"{We:.2f}")

            # 存储校准数据
            self.calib_data = CalibrationData(
                t_raw=t_final_raw,
                F_raw=F_final_raw,
                t_smooth=t_final_smooth,
                F_smooth=F_final_smooth,
                t_rho=t_tau_rho_smooth,
                t_gamma=t_tau_gamma_smooth,
                F_star=F_star_smooth
            )

            # 更新原始数据图
            self.ax_raw.clear()
            self.configure_axes(self.ax_raw, "原始数据曲线", "时间 (s)", "电压 (V)")
            self.ax_raw.plot(self.x, self.y, linewidth=0.8, color='blue', alpha=0.7)

            # 显示当前T0位置
            current_t0 = float(self.peak_info.T0)
            self.ax_raw.axvline(x=current_t0, color='red', linestyle='--', linewidth=0.8, alpha=0.7)
            self.ax_raw.scatter([current_t0], [self.peak_info.V0], s=40,
                                c='red', edgecolors='white', linewidth=0.5, zorder=5)

            # 更新平滑数据图
            self.ax_smooth.clear()
            if self.filter_type == 'gaussian':
                title = "高斯滤波曲线 (在FIR低通滤波后)"
            else:
                title = "FIR低通滤波曲线 - 点击快速定位峰值"
            self.configure_axes(self.ax_smooth, title, "时间 (s)", "电压 (V)")
            self.ax_smooth.plot(self.x, self.y_smoothed_global, linewidth=1.0, color='orange')
            self.ax_smooth.axvline(x=current_t0, color='red', linestyle='--', linewidth=0.8, alpha=0.7)
            self.ax_smooth.scatter([current_t0], [self.peak_info.V0], s=40,
                                   c='red', edgecolors='white', linewidth=0.5, zorder=5)

            # 更新校准图（原始）
            self.ax_calib.clear()
            self.configure_axes(self.ax_calib, "校准曲线（原始）", "相对时间 (s)", "碰撞力 (N)")
            self.ax_calib.plot(t_final_raw, F_final_raw, linewidth=1.0, color='green')
            self.ax_calib.axvline(x=0, color='red', linestyle='--', linewidth=0.8, alpha=0.7)
            self.ax_calib.axhline(y=0, color='blue', linestyle='--', linewidth=0.8, alpha=0.5)

            # 更新校准图（滤波）- 点选参考图
            self.ax_calib_smooth.clear()
            title = "校准曲线（滤波）"
            if self.manual_peak_mode in ['T0', 'Peak1', 'Peak2']:
                title += " - 点击选择点（鼠标悬停显示放大窗口）"
            else:
                title += " - 点击选择点"
            self.configure_axes(self.ax_calib_smooth, title, "相对时间 (s)", "碰撞力 (N)")
            self.ax_calib_smooth.plot(t_final_smooth, F_final_smooth, linewidth=1.2, color='purple')
            self.ax_calib_smooth.axvline(x=0, color='red', linestyle='--', linewidth=0.8, alpha=0.7)
            self.ax_calib_smooth.axhline(y=0, color='blue', linestyle='--', linewidth=0.8, alpha=0.5)

            # 如果有手动峰值，标记它们
            if self.peak_info.is_peak1_manual or self.peak_info.is_peak2_manual:
                self.mark_manual_peaks(t_final_raw, F_final_raw, t_final_smooth, F_final_smooth)

            # 更新合并的无量纲图
            self.ax_dimensionless.clear()

            # 绘制第一组数据（t/τρ） - 使用底部x轴
            line1, = self.ax_dimensionless.plot(t_tau_rho_smooth, F_star_smooth,
                                                linewidth=1.5, color='red', label='t/τρ')

            # 配置底部x轴
            self.ax_dimensionless.set_xlabel("t/τρ", fontsize=9, labelpad=8, color='red')
            self.ax_dimensionless.tick_params(axis='x', labelcolor='red')
            self.ax_dimensionless.set_ylabel("F*", fontsize=9, labelpad=8)

            # 如果顶部坐标轴已存在，先移除
            if self.ax_dimensionless_top is not None:
                self.ax_dimensionless_top.remove()
                self.ax_dimensionless_top = None

            # 创建顶部x轴用于t/τγ
            self.ax_dimensionless_top = self.ax_dimensionless.twiny()

            # 绘制第二组数据（t/τγ） - 使用顶部x轴
            line2, = self.ax_dimensionless_top.plot(t_tau_gamma_smooth, F_star_smooth,
                                                    linewidth=1.5, color='blue', label='t/τγ', alpha=0.8)

            # 配置顶部x轴
            self.ax_dimensionless_top.set_xlabel("t/τγ", fontsize=9, labelpad=8, color='blue')
            self.ax_dimensionless_top.tick_params(axis='x', labelcolor='blue')

            # 设置网格和标题
            self.ax_dimensionless.grid(True, alpha=0.3)
            self.ax_dimensionless.set_title("无量纲曲线", fontsize=10, pad=12)

            # 添加图例
            lines = [line1, line2]
            labels = [line.get_label() for line in lines]
            self.ax_dimensionless.legend(lines, labels, loc='upper right', fontsize=9, framealpha=0.9)

            # 调整布局
            self.fig.tight_layout()
            self.canvas.draw()

        except Exception as e:
            print(f"更新图表时出错: {e}")
            import traceback
            traceback.print_exc()

    def get_calib_data(self, params, T0, V0, delta_t_user, delta_F_user, y_source):
        """获取校准数据"""
        if y_source is None or len(y_source) == 0:
            return np.array([]), np.array([])

        x_min = T0 - params.PRE_TIME
        x_max = T0 + params.POST_TIME

        # 创建插值点
        x_interp = np.arange(x_min, x_max + params.STEP / 2, params.STEP)

        # 提取有效数据
        valid_mask = (self.x >= x_min) & (self.x <= x_max)
        if np.sum(valid_mask) < 5:
            return np.array([]), np.array([])

        x_valid = self.x[valid_mask]
        y_valid = y_source[valid_mask]

        # 插值
        if SCIPY_AVAILABLE and len(x_valid) > 3:
            try:
                interp_func = interp1d(x_valid, y_valid, kind='cubic', fill_value='extrapolate')
                y_interp = interp_func(x_interp)
            except:
                y_interp = np.interp(x_interp, x_valid, y_valid)
        else:
            y_interp = np.interp(x_interp, x_valid, y_valid)

        # 计算相对时间和力
        t_rel_base = x_interp - T0
        F_rel_base = params.COEFF_A * (y_interp - V0)

        # 应用用户偏移
        t_final = t_rel_base - delta_t_user
        F_final = F_rel_base - delta_F_user

        return t_final, F_final

    # ========== 数据重构引擎（用于导出功能）==========
    def reconstruct_curve_data(self, curve_dict):
        """
        从存储的元数据和缓存数据重构完整的曲线数据

        参数:
            curve_dict: 包含metadata和cached_data的曲线字典

        返回:
            包含6种数据类型的字典:
            - dimensionless_trho_original: (t_rho, F_star) 无量纲t/τρ（滤波）
            - dimensionless_tgamma_original: (t_gamma, F_star) 无量纲t/τγ（滤波）
            - filtered_original: (t_smooth, F_smooth) 有量纲（滤波）
            - unfiltered_original: (t_raw, F_raw) 有量纲（未滤波）
            - dimensionless_trho_raw: (t_rho_raw, F_star_raw) 无量纲t/τρ（未滤波）
            - dimensionless_tgamma_raw: (t_gamma_raw, F_star_raw) 无量纲t/τγ（未滤波）
        """
        try:
            metadata = curve_dict["metadata"]
            cached = curve_dict["cached_data"]

            # 提取元数据
            U0 = metadata["U0"]
            D0 = metadata["D0"]
            rho = metadata["rho"]
            sigma = metadata["sigma"]

            # 计算特征时间尺度
            tau_rho = (rho * D0**3) / sigma  # 密度时间尺度
            tau_gamma = np.sqrt((rho * D0**3) / sigma)  # 惯性-表面张力时间尺度

            # 1. 滤波数据（已存储）
            t_smooth = cached["t_smooth"]
            F_smooth = cached["F_smooth"]

            # 2. 未滤波数据（已存储）
            t_raw = cached["t_raw"]
            F_raw = cached["F_raw"]

            # 3. 无量纲数据（滤波）- 直接使用已存储的
            t_rho = curve_dict["t_rho"]
            t_gamma = curve_dict["t_gamma"]
            F_star = curve_dict["F_star"]

            # 4. 无量纲数据（未滤波）- 需要重新计算
            # 计算未滤波的无量纲力
            F_star_raw = F_raw / (sigma * D0)

            # 计算未滤波的无量纲时间
            t_rho_raw = t_raw / tau_rho
            t_gamma_raw = t_raw / tau_gamma

            return {
                "dimensionless_trho_original": (t_rho, F_star),
                "dimensionless_tgamma_original": (t_gamma, F_star),
                "filtered_original": (t_smooth, F_smooth),
                "unfiltered_original": (t_raw, F_raw),
                "dimensionless_trho_raw": (t_rho_raw, F_star_raw),
                "dimensionless_tgamma_raw": (t_gamma_raw, F_star_raw)
            }

        except Exception as e:
            print(f"重构曲线数据时出错: {e}")
            traceback.print_exc()
            return None

    def calculate_unified_ranges(self, curves_data_list, data_type='dimensional'):
        """
        计算统一的X轴范围

        参数:
            curves_data_list: 多条曲线的 (x, y) 数据列表
            data_type: 'dimensionless' 使用并集，'dimensional' 使用交集

        返回:
            (x_min, x_max, num_points) 统一的X轴范围参数
        """
        if not curves_data_list:
            return None

        x_mins = []
        x_maxs = []
        densities = []

        for x_data, y_data in curves_data_list:
            if len(x_data) > 0:
                x_mins.append(np.min(x_data))
                x_maxs.append(np.max(x_data))
                # 计算数据密度（点数/范围）
                x_range = np.max(x_data) - np.min(x_data)
                if x_range > 0:
                    densities.append(len(x_data) / x_range)

        if not x_mins or not x_maxs:
            return None

        # 根据数据类型选择并集或交集
        if data_type == 'dimensionless':
            # 无量纲数据：使用并集（最小起点到最大终点）
            x_min = min(x_mins)
            x_max = max(x_maxs)
        else:
            # 有量纲数据：使用交集（最大起点到最小终点）
            x_min = max(x_mins)
            x_max = min(x_maxs)

        # 检查交集是否有效
        if x_min >= x_max:
            print(f"警告: {data_type} 数据的X轴范围无效 (x_min={x_min}, x_max={x_max})")
            return None

        # 基于最密集曲线的密度计算点数
        if densities:
            max_density = max(densities)
            num_points = int((x_max - x_min) * max_density)
            # 限制点数范围
            num_points = max(100, min(num_points, 50000))
        else:
            num_points = 1000

        return x_min, x_max, num_points

    def interpolate_to_unified_axis(self, curves_data_list, x_unified):
        """
        将多条曲线插值到统一的X轴

        参数:
            curves_data_list: 多条曲线的 (x, y) 数据列表
            x_unified: 统一的X轴数组

        返回:
            插值后的Y值列表，超出范围的点填充NaN
        """
        interpolated_curves = []

        for x_data, y_data in curves_data_list:
            if len(x_data) == 0 or len(y_data) == 0:
                # 空数据，填充全NaN
                interpolated_curves.append(np.full_like(x_unified, np.nan))
                continue

            try:
                # 使用scipy或numpy进行插值
                if SCIPY_AVAILABLE:
                    # scipy的interp1d，超出范围填充NaN
                    f = interp1d(x_data, y_data, kind='linear',
                                bounds_error=False, fill_value=np.nan)
                    y_interp = f(x_unified)
                else:
                    # numpy的interp，手动处理边界
                    y_interp = np.interp(x_unified, x_data, y_data)
                    # 标记超出范围的点为NaN
                    x_min, x_max = np.min(x_data), np.max(x_data)
                    mask = (x_unified < x_min) | (x_unified > x_max)
                    y_interp[mask] = np.nan

                interpolated_curves.append(y_interp)

            except Exception as e:
                print(f"插值时出错: {e}")
                interpolated_curves.append(np.full_like(x_unified, np.nan))

        return interpolated_curves

    def mark_manual_peaks(self, t_raw, F_raw, t_smooth, F_smooth):
        """标记手动峰值"""
        try:
            # 在滤波校准图上标记
            if self.peak_info.is_peak1_manual and self.peak_info.peak1_time_rel is not None:
                p1t = self.peak_info.peak1_time_rel - self.delta_t_user
                p1F_smooth = np.interp(p1t, t_smooth, F_smooth) if len(t_smooth) > 0 else 0
                self.ax_calib_smooth.scatter([p1t], [p1F_smooth],
                                             s=60, c='orange',
                                             edgecolors='white', linewidth=0.5, zorder=5)

            if self.peak_info.is_peak2_manual and self.peak_info.peak2_time_rel is not None:
                p2t = self.peak_info.peak2_time_rel - self.delta_t_user
                p2F_smooth = np.interp(p2t, t_smooth, F_smooth) if len(t_smooth) > 0 else 0
                self.ax_calib_smooth.scatter([p2t], [p2F_smooth],
                                             s=60, c='cyan',
                                             edgecolors='white', linewidth=0.5, zorder=5)

        except Exception as e:
            print(f"标记手动峰值时出错: {e}")

    # ========== 快速定位功能 ==========
    def quick_locate_peak(self, click_time):
        """快速定位点击位置附近的峰值"""
        if len(self.x) == 0 or len(self.y_smoothed_global) == 0:
            return

        try:
            # 在点击位置附近寻找局部最高峰值
            search_window = 0.001  # 搜索窗口大小（秒）
            idx, T0, V0 = self.find_local_peak(
                self.x,
                self.y_smoothed_global,
                click_time,
                search_window
            )

            self.peak_info.is_manual = True
            self.peak_info.idx = idx
            self.peak_info.T0 = T0
            self.peak_info.V0 = V0

            # 重置偏移
            self.delta_t_user = 0.0
            self.delta_F_user = 0.0

            # 更新T0显示
            self.t0_value_label.setText(f"{T0:.6f} s")
            self.quick_locate_info.setText(f"已快速定位: {T0:.6f} s")

            # 更新所有图表
            self.update_all_plots()

            self.statusBar().showMessage(f"已快速定位到峰值: {T0:.6f} s")

        except Exception as e:
            self.statusBar().showMessage(f"快速定位失败: {str(e)}")

    # ========== 放大窗口功能 ==========
    def create_magnifier_window(self):
        """创建放大窗口"""
        if self.magnifier_window is None:
            self.magnifier_window = MagnifierWindow(self)

    def show_magnifier(self, x_pos, label_text="放大视图"):
        """显示放大窗口"""
        if self.calib_data is None:
            return

        if self.magnifier_window is None:
            self.create_magnifier_window()

        # 更新放大窗口内容
        self.magnifier_window.update_magnifier(
            self.calib_data.t_smooth,
            self.calib_data.F_smooth,
            x_pos,
            label_text
        )

        # 移动窗口到右上角
        if not self.magnifier_window.isVisible():
            self.magnifier_window.show()

        # 获取主窗口位置和大小
        main_window_rect = self.geometry()

        # 计算放大窗口位置（右上角）
        window_x = main_window_rect.x() + main_window_rect.width() - self.magnifier_window.width() - 20
        window_y = main_window_rect.y() + 50  # 离主窗口上边缘50像素

        self.magnifier_window.move(window_x, window_y)

    def close_magnifier(self):
        """关闭放大窗口"""
        if self.magnifier_window and self.magnifier_window.isVisible():
            self.magnifier_window.hide()

    def on_mouse_move(self, event):
        """处理鼠标移动事件"""
        # 只有在手动选点模式下才显示放大窗口
        if event.inaxes != self.ax_calib_smooth:
            if not self.adjusting_peak:  # 如果不是正在调整峰值，才关闭放大窗口
                self.close_magnifier()
            return

        x_click = event.xdata
        y_click = event.ydata

        if x_click is None or y_click is None:
            if not self.adjusting_peak:  # 如果不是正在调整峰值，才关闭放大窗口
                self.close_magnifier()
            return

        # 在手动选点模式下显示放大窗口
        if self.manual_peak_mode in ['T0', 'Peak1', 'Peak2'] and self.magnifier_active:
            if self.manual_peak_mode == 'T0':
                label_text = "T0校准放大视图"
            elif self.manual_peak_mode == 'Peak1':
                label_text = "Peak1选择放大视图"
            elif self.manual_peak_mode == 'Peak2':
                label_text = "Peak2选择放大视图"
            else:
                label_text = "放大视图"

            self.show_magnifier(x_click, label_text)

    # ========== T0微调功能 ==========
    def adjust_t0(self, direction):
        """微调T0（带放大窗口）"""
        if self.peak_info is None or self.peak_info.T0 is None:
            return

        try:
            # 标记正在调整
            self.adjusting_peak = True

            # 计算调整量
            adjustment = direction * self.t0_adjust_step
            new_T0 = self.peak_info.T0 + adjustment

            # 找到最近的索引
            idx = int(np.argmin(np.abs(self.x - new_T0)))

            if idx < 5 or idx > len(self.x) - 5:
                self.statusBar().showMessage("警告：调整后的T0接近数据边缘")
                self.adjusting_peak = False
                return

            # 更新T0
            self.peak_info.is_manual = True
            self.peak_info.idx = idx
            self.peak_info.T0 = float(self.x[idx])
            self.peak_info.V0 = float(self.y_smoothed_global[idx])

            # 重置偏移
            self.delta_t_user = 0.0
            self.delta_F_user = 0.0

            # 更新显示
            self.t0_value_label.setText(f"{self.peak_info.T0:.6f} s")
            self.quick_locate_info.setText(f"已微调T0: {self.peak_info.T0:.6f} s")

            # 更新图表
            self.update_all_plots()

            # 显示放大窗口（显示当前零点位置）
            if self.calib_data is not None:
                # 计算零点在相对时间坐标中的位置
                zero_point = 0.0  # 校准后的零点就是0
                self.show_magnifier(zero_point, f"T0微调: {direction} {self.t0_adjust_step:.7f} s")

            direction_text = "左移" if direction < 0 else "右移"
            self.statusBar().showMessage(f"已微调T0: {direction_text} {self.t0_adjust_step:.7f} s")

            # 延迟重置调整标志
            QTimer.singleShot(500, lambda: setattr(self, 'adjusting_peak', False))

        except Exception as e:
            self.statusBar().showMessage(f"微调失败: {str(e)}")
            self.adjusting_peak = False

    # ========== 峰值微调功能（带放大窗口） ==========
    def adjust_peak_time_with_magnifier(self, peak_type, direction):
        """微调峰值时间（带放大窗口）"""
        try:
            # 标记正在调整
            self.adjusting_peak = True

            if peak_type == 'Peak1' and self.peak_info.is_peak1_manual and self.peak_info.peak1_time_rel is not None:
                # 微调Peak1
                current_time = self.peak_info.peak1_time_rel - self.delta_t_user
                adjustment = direction * self.peak_adjust_step
                new_time = current_time + adjustment

                # 更新峰值时间
                self.peak_info.peak1_time_rel = new_time + self.delta_t_user
                self.peak1_edit.setText(f"{new_time:.6f}")

                # 更新显示和图表
                self.apply_manual_peaks()

                # 显示放大窗口
                if self.calib_data is not None:
                    self.show_magnifier(new_time, f"Peak1微调: {direction} {self.peak_adjust_step:.7f} s")

                direction_text = "左移" if direction < 0 else "右移"
                self.statusBar().showMessage(f"已微调Peak1: {direction_text} {self.peak_adjust_step:.7f} s")

            elif peak_type == 'Peak2' and self.peak_info.is_peak2_manual and self.peak_info.peak2_time_rel is not None:
                # 微调Peak2
                current_time = self.peak_info.peak2_time_rel - self.delta_t_user
                adjustment = direction * self.peak_adjust_step
                new_time = current_time + adjustment

                # 更新峰值时间
                self.peak_info.peak2_time_rel = new_time + self.delta_t_user
                self.peak2_edit.setText(f"{new_time:.6f}")

                # 更新显示和图表
                self.apply_manual_peaks()

                # 显示放大窗口
                if self.calib_data is not None:
                    self.show_magnifier(new_time, f"Peak2微调: {direction} {self.peak_adjust_step:.7f} s")

                direction_text = "左移" if direction < 0 else "右移"
                self.statusBar().showMessage(f"已微调Peak2: {direction_text} {self.peak_adjust_step:.7f} s")
            else:
                self.statusBar().showMessage(f"请先设置{peak_type}峰值")

            # 延迟重置调整标志
            QTimer.singleShot(500, lambda: setattr(self, 'adjusting_peak', False))

        except Exception as e:
            self.statusBar().showMessage(f"微调失败: {str(e)}")
            self.adjusting_peak = False

    # ========== T0校准功能 ==========
    def start_t0_calibration(self):
        """开始T0校准"""
        if len(self.x) == 0 or len(self.y_smoothed_global) == 0:
            QMessageBox.warning(self, "提示", "请先加载数据")
            return

        self.manual_peak_mode = 'T0'
        self.magnifier_active = True
        self.calibrate_t0_btn.setEnabled(False)
        self.cancel_calibrate_btn.setEnabled(True)
        self.calibration_instruction.setText(
            "请在'校准曲线（滤波）'图中点击选择零点位置\n"
            "鼠标悬停时会显示放大窗口，便于精确选择\n"
            "或者使用微调按钮进行精细调整（微调时会显示放大窗口）"
        )

        self.statusBar().showMessage("T0校准模式：在校准曲线图中点击选择零点位置，鼠标悬停时显示放大窗口，或使用微调按钮")

    def cancel_calibration(self):
        """取消校准"""
        self.manual_peak_mode = None
        self.magnifier_active = False
        self.adjusting_peak = False
        self.calibrate_t0_btn.setEnabled(True)
        self.cancel_calibrate_btn.setEnabled(False)
        self.calibration_instruction.setText(
            "提示：点击'精确校准T0'按钮后，在'校准曲线（滤波）'图中点击选择零点位置\n"
            "鼠标悬停时会显示放大窗口，便于精确选择\n"
            "或者使用微调按钮进行精细调整（微调时会显示放大窗口）"
        )

        # 关闭放大窗口
        self.close_magnifier()

        self.statusBar().showMessage("已取消校准模式")

    # ========== 手动峰值操作 ==========
    def start_manual_peak_picking(self, mode):
        """开始手动选点模式"""
        if len(self.x) == 0 or len(self.y_smoothed_global) == 0:
            QMessageBox.warning(self, "提示", "请先加载数据")
            return

        if self.peak_info.T0 is None:
            QMessageBox.warning(self, "提示", "请先设置T0")
            return

        self.manual_peak_mode = mode
        self.magnifier_active = True
        self.adjusting_peak = False

        # 启用峰值微调按钮（只在已经设置了峰值时才启用）
        if mode == 'Peak1' and self.peak_info.is_peak1_manual:
            self.peak1_left_btn.setEnabled(True)
            self.peak1_right_btn.setEnabled(True)
        elif mode == 'Peak2' and self.peak_info.is_peak2_manual:
            self.peak2_left_btn.setEnabled(True)
            self.peak2_right_btn.setEnabled(True)

        self.statusBar().showMessage(
            f"{mode}选择模式：在校准曲线图中点击选择峰值位置，鼠标悬停时显示放大窗口，或使用微调按钮")

    def auto_detect_peak(self, mode):
        """自动检测峰值"""
        if self.calib_data is None:
            self.statusBar().showMessage("请先加载数据并完成校准")
            return

        try:
            if mode == 'Peak1':
                positive_mask = self.calib_data.t_smooth > 0
                if np.sum(positive_mask) == 0:
                    return

                t_positive = self.calib_data.t_smooth[positive_mask]
                F_positive = self.calib_data.F_smooth[positive_mask]

                dF = np.diff(F_positive)
                peak_indices = np.where((dF[:-1] > 0) & (dF[1:] < 0))[0] + 1

                if len(peak_indices) > 0:
                    peak_idx = peak_indices[0]
                    peak_time = t_positive[peak_idx]
                    self.peak1_edit.setText(f"{peak_time:.6f}")
                    self.apply_manual_peaks()

                    # 启用微调按钮
                    self.peak1_left_btn.setEnabled(True)
                    self.peak1_right_btn.setEnabled(True)

            elif mode == 'Peak2':
                positive_mask = self.calib_data.t_smooth > 0
                if np.sum(positive_mask) == 0:
                    return

                t_positive = self.calib_data.t_smooth[positive_mask]
                F_positive = self.calib_data.F_smooth[positive_mask]

                dF = np.diff(F_positive)
                peak_indices = np.where((dF[:-1] > 0) & (dF[1:] < 0))[0] + 1

                if len(peak_indices) > 1:
                    peak_idx = peak_indices[1] if len(peak_indices) > 1 else peak_indices[0]
                    peak_time = t_positive[peak_idx]
                    self.peak2_edit.setText(f"{peak_time:.6f}")
                    self.apply_manual_peaks()

                    # 启用微调按钮
                    self.peak2_left_btn.setEnabled(True)
                    self.peak2_right_btn.setEnabled(True)

        except Exception as e:
            self.statusBar().showMessage(f"自动检测失败: {str(e)}")

    def on_canvas_click(self, event):
        """处理画布点击事件"""
        if event.inaxes is None:
            return

        x_click = event.xdata
        y_click = event.ydata

        if x_click is None or y_click is None:
            return

        # 处理FIR低通滤波曲线图中的快速定位（不在任何手动模式下）
        if event.inaxes == self.ax_smooth and self.manual_peak_mode is None:
            self.quick_locate_peak(x_click)
            return

        if self.manual_peak_mode is None:
            return

        if self.manual_peak_mode == 'T0':
            # T0校准：在校准滤波曲线图中点击
            if event.inaxes == self.ax_calib_smooth:
                # 点击位置是相对时间，需要转换为新的T0
                t_rel = x_click  # 相对时间

                # 计算新的T0：当前T0 + 相对时间 - delta_t_user
                new_T0 = float(self.peak_info.T0) + t_rel - self.delta_t_user

                # 找到最近的数据点
                idx = int(np.argmin(np.abs(self.x - new_T0)))

                if idx < 5 or idx > len(self.x) - 5:
                    self.statusBar().showMessage("警告：选择的T0接近数据边缘")
                    return

                self.peak_info.is_manual = True
                self.peak_info.idx = idx
                self.peak_info.T0 = float(self.x[idx])
                self.peak_info.V0 = float(self.y_smoothed_global[idx])

                # 重置偏移
                self.delta_t_user = 0.0
                self.delta_F_user = 0.0

                # 更新T0显示
                self.t0_value_label.setText(f"{self.peak_info.T0:.6f} s")
                self.quick_locate_info.setText(f"已精确校准: {self.peak_info.T0:.6f} s")

                # 退出校准模式
                self.manual_peak_mode = None
                self.magnifier_active = False
                self.adjusting_peak = False
                self.calibrate_t0_btn.setEnabled(True)
                self.cancel_calibrate_btn.setEnabled(False)
                self.calibration_instruction.setText(
                    "提示：点击'精确校准T0'按钮后，在'校准曲线（滤波）'图中点击选择零点位置\n"
                    "鼠标悬停时会显示放大窗口，便于精确选择\n"
                    "或者使用微调按钮进行精细调整（微调时会显示放大窗口）"
                )

                # 关闭放大窗口
                self.close_magnifier()

                # 更新所有图表
                self.update_all_plots()

                self.statusBar().showMessage(f"已校准T0: {self.peak_info.T0:.6f} s")

        elif self.manual_peak_mode == 'Peak1':
            # Peak1选择：在校准滤波曲线图中选择
            if event.inaxes == self.ax_calib_smooth:
                self.peak1_edit.setText(f"{x_click:.6f}")

                # 退出选点模式
                self.manual_peak_mode = None
                self.magnifier_active = False
                self.adjusting_peak = False

                # 关闭放大窗口
                self.close_magnifier()

                self.statusBar().showMessage("Peak1已选择")
                self.apply_manual_peaks()

        elif self.manual_peak_mode == 'Peak2':
            # Peak2选择：在校准滤波曲线图中选择
            if event.inaxes == self.ax_calib_smooth:
                self.peak2_edit.setText(f"{x_click:.6f}")

                # 退出选点模式
                self.manual_peak_mode = None
                self.magnifier_active = False
                self.adjusting_peak = False

                # 关闭放大窗口
                self.close_magnifier()

                self.statusBar().showMessage("Peak2已选择")
                self.apply_manual_peaks()

    def apply_manual_peaks(self):
        """应用手动峰值"""
        p1_str = self.peak1_edit.text().strip()
        p2_str = self.peak2_edit.text().strip()

        if not p1_str and not p2_str:
            self.statusBar().showMessage("请输入至少一个峰值的时间")
            return

        try:
            if p1_str:
                peak1_time_calib = float(p1_str)
                peak1_time_rel = peak1_time_calib + self.delta_t_user
                self.peak_info.is_peak1_manual = True
                self.peak_info.peak1_time_rel = float(peak1_time_rel)

                # 启用Peak1微调按钮
                self.peak1_left_btn.setEnabled(True)
                self.peak1_right_btn.setEnabled(True)

            if p2_str:
                peak2_time_calib = float(p2_str)
                peak2_time_rel = peak2_time_calib + self.delta_t_user
                self.peak_info.is_peak2_manual = True
                self.peak_info.peak2_time_rel = float(peak2_time_rel)

                # 启用Peak2微调按钮
                self.peak2_left_btn.setEnabled(True)
                self.peak2_right_btn.setEnabled(True)

            self.update_all_plots()
            self.extract_manual_peaks()

            self.statusBar().showMessage("已应用手动峰值")

        except Exception as e:
            self.statusBar().showMessage(f"设置失败: {str(e)}")

    def extract_manual_peaks(self):
        """提取手动峰值数据"""
        if self.calib_data is None:
            return

        try:
            # Peak1
            if self.peak_info.is_peak1_manual and self.peak_info.peak1_time_rel is not None:
                p1t = self.peak_info.peak1_time_rel - self.delta_t_user
                t_smooth = self.calib_data.t_smooth
                if len(t_smooth) > 0:
                    p1_idx = int(np.argmin(np.abs(t_smooth - p1t)))
                    self.peak1_trho_edit.setText(f"{self.calib_data.t_rho[p1_idx]:.6f}")
                    self.peak1_tgamma_edit.setText(f"{self.calib_data.t_gamma[p1_idx]:.6f}")
                    self.peak1_fstar_edit.setText(f"{self.calib_data.F_star[p1_idx]:.6f}")

            # Peak2
            if self.peak_info.is_peak2_manual and self.peak_info.peak2_time_rel is not None:
                p2t = self.peak_info.peak2_time_rel - self.delta_t_user
                t_smooth = self.calib_data.t_smooth
                if len(t_smooth) > 0:
                    p2_idx = int(np.argmin(np.abs(t_smooth - p2t)))
                    self.peak2_trho_edit.setText(f"{self.calib_data.t_rho[p2_idx]:.6f}")
                    self.peak2_tgamma_edit.setText(f"{self.calib_data.t_gamma[p2_idx]:.6f}")
                    self.peak2_fstar_edit.setText(f"{self.calib_data.F_star[p2_idx]:.6f}")

        except Exception as e:
            print(f"提取手动峰值时出错: {e}")

    def extract_peaks_callback(self):
        """自动提取峰值"""
        if self.calib_data is None:
            self.statusBar().showMessage("请先加载数据并完成校准")
            return

        try:
            peaks = self.extract_peaks(self.calib_data.t_rho, self.calib_data.t_gamma, self.calib_data.F_star)

            self.peak1_trho_edit.setText(f"{peaks['peak1']['t_rho']:.6f}")
            self.peak1_tgamma_edit.setText(f"{peaks['peak1']['t_gamma']:.6f}")
            self.peak1_fstar_edit.setText(f"{peaks['peak1']['F_star']:.6f}")

            self.peak2_trho_edit.setText(f"{peaks['peak2']['t_rho']:.6f}")
            self.peak2_tgamma_edit.setText(f"{peaks['peak2']['t_gamma']:.6f}")
            self.peak2_fstar_edit.setText(f"{peaks['peak2']['F_star']:.6f}")

            if peaks['peak1']['t_rel'] is not None:
                self.peak1_edit.setText(f"{peaks['peak1']['t_rel']:.6f}")
                self.peak_info.is_peak1_manual = True
                self.peak_info.peak1_time_rel = float(peaks['peak1']['t_rel'] + self.delta_t_user)
                self.peak1_left_btn.setEnabled(True)
                self.peak1_right_btn.setEnabled(True)

            if peaks['peak2']['t_rel'] is not None:
                self.peak2_edit.setText(f"{peaks['peak2']['t_rel']:.6f}")
                self.peak_info.is_peak2_manual = True
                self.peak_info.peak2_time_rel = float(peaks['peak2']['t_rel'] + self.delta_t_user)
                self.peak2_left_btn.setEnabled(True)
                self.peak2_right_btn.setEnabled(True)

            self.update_all_plots()
            self.statusBar().showMessage("已自动提取峰值")

        except Exception as e:
            self.statusBar().showMessage(f"提取失败: {str(e)}")

    def extract_peaks(self, t_rho, t_gamma, F_star):
        """自动提取峰值算法"""
        if len(F_star) < 3:
            raise ValueError("数据点不足，无法识别极值点")

        positive_mask = self.calib_data.t_smooth > 0
        if np.sum(positive_mask) == 0:
            raise ValueError("没有正时间数据")

        t_positive = self.calib_data.t_smooth[positive_mask]
        F_positive = F_star[positive_mask]

        dF = np.diff(F_positive)
        peak_locs = np.where((dF[:-1] > 0) & (dF[1:] < 0))[0] + 1

        if len(peak_locs) == 0:
            peak_locs = [np.argmax(F_positive)]

        peak_F = F_positive[peak_locs]
        if len(peak_locs) >= 2:
            # 按峰值高度排序
            idx_sorted = np.argsort(-peak_F)[:2]
            sorted_locs = peak_locs[idx_sorted]
            sorted_F = peak_F[idx_sorted]
        else:
            sorted_locs = [peak_locs[0], peak_locs[0]]
            sorted_F = [peak_F[0], peak_F[0]]

        t_rel1 = t_positive[sorted_locs[0]]
        t_rel2 = t_positive[sorted_locs[1]]

        return {
            'peak1': {
                'F_star': float(sorted_F[0]),
                't_rho': float(t_rho[positive_mask][sorted_locs[0]]),
                't_gamma': float(t_gamma[positive_mask][sorted_locs[0]]),
                't_rel': float(t_rel1)
            },
            'peak2': {
                'F_star': float(sorted_F[1]),
                't_rho': float(t_rho[positive_mask][sorted_locs[1]]),
                't_gamma': float(t_gamma[positive_mask][sorted_locs[1]]),
                't_rel': float(t_rel2)
            }
        }

    # ========== 数据导出功能（优化版） ==========
    def export_all_data(self):
        """导出所有数据到一个Excel文件（优化顺序）"""
        if self.calib_data is None:
            self.statusBar().showMessage("请先加载数据并完成校准")
            return

        try:
            fluid_name = self.fluid_name_combo.currentText()
            D0_str = f"{self.D0:.3f}"
            U0_str = f"{self.U0:.3f}"
            default_filename = f"{fluid_name}_{D0_str}_{U0_str}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存所有数据",
                default_filename,
                "Excel文件 (*.xlsx);;所有文件 (*.*)"
            )

            if not file_path:
                return

            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'

            workbook = Workbook()

            # 1. 滤波校准曲线（滤波后的数据）
            ws1 = workbook.active
            ws1.title = "滤波校准曲线"
            ws1.append(["相对时间(s)", "碰撞力(N)"])
            for t, F in zip(self.calib_data.t_smooth, self.calib_data.F_smooth):
                ws1.append([t, F])

            # 2. 无量纲曲线_τρ
            ws2 = workbook.create_sheet(title="无量纲曲线_τρ")
            ws2.append(["t/τρ", "F*"])
            for t, F in zip(self.calib_data.t_rho, self.calib_data.F_star):
                ws2.append([t, F])

            # 3. 无量纲曲线_τγ
            ws3 = workbook.create_sheet(title="无量纲曲线_τγ")
            ws3.append(["t/τγ", "F*"])
            for t, F in zip(self.calib_data.t_gamma, self.calib_data.F_star):
                ws3.append([t, F])

            # 4. 原始校准曲线
            ws4 = workbook.create_sheet(title="原始校准曲线")
            ws4.append(["相对时间(s)", "碰撞力(N)"])
            for t, F in zip(self.calib_data.t_raw, self.calib_data.F_raw):
                ws4.append([t, F])

            workbook.save(file_path)
            self.statusBar().showMessage(f"所有数据已导出到: {file_path}")

        except Exception as e:
            self.statusBar().showMessage(f"导出失败: {str(e)}")

    def export_peaks_callback(self):
        """导出峰值数据到汇总文件"""
        try:
            if (not self.peak1_trho_edit.text() or not self.peak2_trho_edit.text() or
                    not self.weber_edit.text()):
                self.statusBar().showMessage("请先提取峰值数据")
                return

            # 默认导出到当前文件所在目录
            if self.data_folder and os.path.exists(self.data_folder):
                default_dir = self.data_folder
            elif self.current_file_path and os.path.exists(os.path.dirname(self.current_file_path)):
                default_dir = os.path.dirname(self.current_file_path)
            else:
                default_dir = os.getcwd()
            excel_path = os.path.join(default_dir, "韦伯数-无量纲力&时间汇总.xlsx")
            current_time = datetime.now().strftime("%Y-%m-d %H:%M:%S")

            new_data = [
                float(self.weber_edit.text()),
                float(self.u0_edit.text()),
                float(self.d0_edit.text()),
                float(self.peak1_fstar_edit.text()),
                float(self.peak1_trho_edit.text()),
                float(self.peak1_tgamma_edit.text()),
                float(self.peak2_fstar_edit.text()),
                float(self.peak2_trho_edit.text()),
                float(self.peak2_tgamma_edit.text()),
                self.fluid_name_combo.currentText(),
                current_time
            ]

            self.append_to_excel(excel_path, new_data)
            self.statusBar().showMessage(f"峰值数据已成功追加至: {excel_path}")

        except Exception as e:
            self.statusBar().showMessage(f"导出失败: {str(e)}")

    def append_to_excel(self, excel_path, row_data):
        """追加数据到Excel文件"""
        headers = ['We', 'U0(m/s)', 'D0(mm)', 'F1*', 't1/τρ', 't1/τγ', 'F2*', 't2/τρ', 't2/τγ', 'fluid', 'time']

        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            max_row = worksheet.max_row
            for col, value in enumerate(row_data, start=1):
                worksheet.cell(row=max_row + 1, column=col, value=value)
        else:
            workbook = Workbook()
            worksheet = workbook.active
            for col, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col, value=header)
            for col, value in enumerate(row_data, start=1):
                worksheet.cell(row=2, column=col, value=value)

        workbook.save(excel_path)

    def clear_peak_displays(self):
        """清除峰值显示"""
        self.peak1_trho_edit.clear()
        self.peak1_tgamma_edit.clear()
        self.peak1_fstar_edit.clear()
        self.peak2_trho_edit.clear()
        self.peak2_tgamma_edit.clear()
        self.peak2_fstar_edit.clear()

    # ========== 曲线对比功能 ==========
    def add_to_compare(self):
        """将当前标定曲线添加到对比图（增强版：存储完整元数据和缓存数据）"""
        if self.calib_data is None:
            QMessageBox.warning(self, "警告", "请先完成曲线标定")
            return

        # 生成图例标签：流体类型_直径_速度_序号
        label = f"{self.current_fluid_type}_{self.D0}_{self.U0}_{len(self.compare_curves)+1}"

        # 获取当前流体参数
        try:
            rho = float(self.rho_edit.text())
            sigma = float(self.sigma_edit_fluid.text())
        except ValueError:
            QMessageBox.warning(self, "警告", "流体参数无效")
            return

        # 存储曲线数据（包含完整元数据和缓存数据）
        self.compare_curves.append({
            "label": label,
            "t_rho": self.calib_data.t_rho.copy(),
            "t_gamma": self.calib_data.t_gamma.copy(),
            "F_star": self.calib_data.F_star.copy(),
            "visible": True,
            "metadata": {
                "file_path": self.current_file_path,
                "U0": self.U0,
                "D0": self.D0,
                "rho": rho,
                "sigma": sigma,
                "T0": self.peak_info.T0,
                "V0": self.peak_info.V0,
                "delta_t_user": self.delta_t_user,
                "delta_F_user": self.delta_F_user,
                "filter_type": self.filter_type,
                "fir_fc": self.params.FIR_FC,
                "sigma_gauss": self.params.SIGMA,
                "COEFF_A": self.params.COEFF_A,
                "PRE_TIME": self.params.PRE_TIME,
                "POST_TIME": self.params.POST_TIME,
                "STEP": self.params.STEP
            },
            "cached_data": {
                "t_raw": self.calib_data.t_raw.copy(),
                "F_raw": self.calib_data.F_raw.copy(),
                "t_smooth": self.calib_data.t_smooth.copy(),
                "F_smooth": self.calib_data.F_smooth.copy()
            }
        })

        # 更新曲线列表UI、对比图和计数标签
        self.update_curve_list_ui()
        self.update_compare_plot()
        self.compare_count_label.setText(f"已添加: {len(self.compare_curves)} 条曲线")
        self.statusBar().showMessage(f"已添加曲线: {label}")

    def update_compare_plot(self):
        """更新对比曲线图"""
        self.ax_compare.clear()
        self.ax_compare.set_title("对比曲线", fontsize=10, pad=12)
        self.ax_compare.set_xlabel("t/τρ", fontsize=9, labelpad=8)
        self.ax_compare.set_ylabel("F*", fontsize=9, labelpad=8)
        self.ax_compare.grid(True, alpha=0.3)
        self.ax_compare.tick_params(labelsize=8)

        # 只绑定可见的曲线
        has_visible = False
        for curve in self.compare_curves:
            if curve.get("visible", True):
                self.ax_compare.plot(curve["t_rho"], curve["F_star"],
                                    linewidth=1.2, label=curve["label"])
                has_visible = True

        if has_visible:
            self.ax_compare.legend(fontsize=7, loc='best')

        self.canvas.draw_idle()

    def clear_compare(self):
        """清空对比曲线"""
        self.compare_curves.clear()
        self.update_curve_list_ui()
        self.update_compare_plot()
        self.compare_count_label.setText("已添加: 0 条曲线")
        self.statusBar().showMessage("已清空对比图")

    def update_curve_list_ui(self):
        """刷新曲线列表UI"""
        # 清空现有列表项（保留stretch）
        while self.curve_list_layout.count() > 1:
            item = self.curve_list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # 为每条曲线创建一行控件
        curve_count = len(self.compare_curves)
        for idx, curve in enumerate(self.compare_curves):
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(4)

            # 复选框（显示/隐藏）
            checkbox = QCheckBox(curve["label"])
            checkbox.setChecked(curve.get("visible", True))
            checkbox.setStyleSheet("font-size: 8pt;")
            checkbox.stateChanged.connect(lambda state, i=idx: self.toggle_curve_visibility(i, state))
            row_layout.addWidget(checkbox, 1)

            # 上移按钮
            up_btn = QPushButton("↑")
            up_btn.setFixedSize(20, 20)
            up_btn.setStyleSheet("font-size: 10pt;")
            up_btn.clicked.connect(lambda checked, i=idx: self.move_curve_up(i))
            up_btn.setEnabled(idx > 0)  # 第一条曲线禁用上移
            row_layout.addWidget(up_btn)

            # 下移按钮
            down_btn = QPushButton("↓")
            down_btn.setFixedSize(20, 20)
            down_btn.setStyleSheet("font-size: 10pt;")
            down_btn.clicked.connect(lambda checked, i=idx: self.move_curve_down(i))
            down_btn.setEnabled(idx < curve_count - 1)  # 最后一条曲线禁用下移
            row_layout.addWidget(down_btn)

            # 删除按钮
            del_btn = QPushButton("×")
            del_btn.setFixedSize(20, 20)
            del_btn.setStyleSheet("font-size: 10pt; font-weight: bold; color: #c00;")
            del_btn.clicked.connect(lambda checked, i=idx: self.remove_curve(i))
            row_layout.addWidget(del_btn)

            self.curve_list_layout.insertWidget(idx, row_widget)

    def toggle_curve_visibility(self, index, state):
        """切换指定曲线的可见性"""
        if 0 <= index < len(self.compare_curves):
            # state: 0=Qt.Unchecked, 2=Qt.Checked
            self.compare_curves[index]["visible"] = (state == 2)
            self.update_compare_plot()

    def move_curve_up(self, index):
        """将曲线上移一位"""
        if index > 0 and index < len(self.compare_curves):
            # 交换位置
            self.compare_curves[index], self.compare_curves[index-1] = \
                self.compare_curves[index-1], self.compare_curves[index]
            self.update_curve_list_ui()
            self.update_compare_plot()

    def move_curve_down(self, index):
        """将曲线下移一位"""
        if index >= 0 and index < len(self.compare_curves) - 1:
            # 交换位置
            self.compare_curves[index], self.compare_curves[index+1] = \
                self.compare_curves[index+1], self.compare_curves[index]
            self.update_curve_list_ui()
            self.update_compare_plot()

    def remove_curve(self, index):
        """删除指定曲线"""
        if 0 <= index < len(self.compare_curves):
            removed_label = self.compare_curves[index]["label"]
            del self.compare_curves[index]
            self.update_curve_list_ui()
            self.update_compare_plot()
            self.compare_count_label.setText(f"已添加: {len(self.compare_curves)} 条曲线")
            self.statusBar().showMessage(f"已删除曲线: {removed_label}")

    def export_compare_data(self):
        """导出对比图数据到Excel（增强版：8个Sheet）"""
        if not self.compare_curves:
            QMessageBox.warning(self, "警告",
                              "对比图中没有曲线数据\n\n"
                              "操作步骤：\n"
                              "1. 加载并标定曲线\n"
                              "2. 点击「添加到对比」按钮\n"
                              "3. 重复步骤1-2添加更多曲线\n"
                              "4. 点击「导出」按钮")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "警告", "openpyxl未安装，无法导出Excel")
            return

        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出对比数据（8个Sheet）",
            os.path.join(self.data_folder or "", "compare_data_8sheets.xlsx") if self.data_folder else "compare_data_8sheets.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return

        try:
            # 重构所有曲线的数据
            all_curves_data = []
            for curve in self.compare_curves:
                reconstructed = self.reconstruct_curve_data(curve)
                if reconstructed is None:
                    QMessageBox.critical(self, "错误", f"重构曲线数据失败: {curve['label']}")
                    return
                all_curves_data.append({
                    "label": curve["label"],
                    "data": reconstructed
                })

            workbook = Workbook()
            # 删除默认的Sheet
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]

            # Sheet 1: t/τρ原始数据（无量纲，滤波）
            self._export_sheet_original(
                workbook, "1_t_τρ原始数据",
                all_curves_data, "dimensionless_trho_original",
                "t/τρ", "F*"
            )

            # Sheet 2: t/τρ插值对齐数据（无量纲，滤波）
            self._export_sheet_interpolated(
                workbook, "2_t_τρ插值对齐",
                all_curves_data, "dimensionless_trho_original",
                "t/τρ", "dimensionless"
            )

            # Sheet 3: t/τγ原始数据（无量纲，滤波）
            self._export_sheet_original(
                workbook, "3_t_τγ原始数据",
                all_curves_data, "dimensionless_tgamma_original",
                "t/τγ", "F*"
            )

            # Sheet 4: t/τγ插值对齐数据（无量纲，滤波）
            self._export_sheet_interpolated(
                workbook, "4_t_τγ插值对齐",
                all_curves_data, "dimensionless_tgamma_original",
                "t/τγ", "dimensionless"
            )

            # Sheet 5: 滤波后原始数据（有量纲）
            self._export_sheet_original(
                workbook, "5_滤波后原始数据",
                all_curves_data, "filtered_original",
                "t (s)", "F (N)"
            )

            # Sheet 6: 滤波后插值对齐数据（有量纲）
            self._export_sheet_interpolated(
                workbook, "6_滤波后插值对齐",
                all_curves_data, "filtered_original",
                "t (s)", "dimensional"
            )

            # Sheet 7: 未滤波原始数据（有量纲）
            self._export_sheet_original(
                workbook, "7_未滤波原始数据",
                all_curves_data, "unfiltered_original",
                "t (s)", "F (N)"
            )

            # Sheet 8: 未滤波插值对齐数据（有量纲）
            self._export_sheet_interpolated(
                workbook, "8_未滤波插值对齐",
                all_curves_data, "unfiltered_original",
                "t (s)", "dimensional"
            )

            workbook.save(file_path)
            self.statusBar().showMessage(f"已导出对比数据（8个Sheet）: {file_path}")
            QMessageBox.information(self, "成功",
                                  f"对比数据已导出到:\n{file_path}\n\n"
                                  f"包含8个Sheet:\n"
                                  f"1. t/τρ原始数据（无量纲，滤波）\n"
                                  f"2. t/τρ插值对齐（并集范围）\n"
                                  f"3. t/τγ原始数据（无量纲，滤波）\n"
                                  f"4. t/τγ插值对齐（并集范围）\n"
                                  f"5. 滤波后原始数据（有量纲）\n"
                                  f"6. 滤波后插值对齐（交集范围）\n"
                                  f"7. 未滤波原始数据（有量纲）\n"
                                  f"8. 未滤波插值对齐（交集范围）")

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}\n\n{traceback.format_exc()}")

    def _export_sheet_original(self, workbook, sheet_name, all_curves_data, data_key, x_label, y_label):
        """
        导出原始数据（不插值）

        表格格式：| X_curve1 | Y_curve1 | X_curve2 | Y_curve2 | ...
        """
        ws = workbook.create_sheet(sheet_name)

        # 构建表头：每条曲线两列
        header = []
        for curve_info in all_curves_data:
            label = curve_info["label"]
            header.extend([f"{x_label}_{label}", f"{y_label}_{label}"])
        ws.append(header)

        # 找出最长的曲线
        max_len = 0
        for curve_info in all_curves_data:
            x_data, y_data = curve_info["data"][data_key]
            max_len = max(max_len, len(x_data))

        # 写入数据
        for i in range(max_len):
            row = []
            for curve_info in all_curves_data:
                x_data, y_data = curve_info["data"][data_key]
                if i < len(x_data):
                    row.extend([x_data[i], y_data[i]])
                else:
                    row.extend(["", ""])
            ws.append(row)

    def _export_sheet_interpolated(self, workbook, sheet_name, all_curves_data, data_key, x_label, data_type):
        """
        导出插值对齐数据

        表格格式：| X_unified | Y_curve1 | Y_curve2 | Y_curve3 | ...
        data_type: 'dimensionless' 使用并集，'dimensional' 使用交集
        """
        ws = workbook.create_sheet(sheet_name)

        # 提取所有曲线的数据
        curves_data_list = []
        for curve_info in all_curves_data:
            x_data, y_data = curve_info["data"][data_key]
            curves_data_list.append((x_data, y_data))

        # 计算统一的X轴范围
        range_params = self.calculate_unified_ranges(curves_data_list, data_type)
        if range_params is None:
            # 范围无效，写入警告信息
            ws.append([f"警告: {data_type} 数据的X轴范围无效，无法生成插值对齐数据"])
            return

        x_min, x_max, num_points = range_params
        x_unified = np.linspace(x_min, x_max, num_points)

        # 插值到统一X轴
        interpolated_curves = self.interpolate_to_unified_axis(curves_data_list, x_unified)

        # 构建表头（根据数据类型判断使用F*还是F）
        if data_key.startswith("dimensionless"):
            # 无量纲数据使用F*
            header = [x_label] + [f"F*_{curve_info['label']}" for curve_info in all_curves_data]
        else:
            # 有量纲数据使用F
            header = [x_label] + [f"F_{curve_info['label']}" for curve_info in all_curves_data]
        ws.append(header)

        # 写入数据
        for i, x in enumerate(x_unified):
            row = [x] + [y_interp[i] if not np.isnan(y_interp[i]) else ""
                        for y_interp in interpolated_curves]
            ws.append(row)


# ========== 主程序 ==========
def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    window = PeakAnalysisTool()
    window.show()

    sys.exit(app.exec())


if __name__ == '__main__':
    main()