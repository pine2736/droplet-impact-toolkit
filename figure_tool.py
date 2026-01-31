# -*- coding: utf-8 -*-
"""
Excel Multi-Sheet Log-Log Plotter (Plotly + Dash)
+ Integrated Non-dimensional Parameter Updater (per-sheet 6-params)

New in this integrated version
1) Excel path is NOT hard-coded: prompt at runtime if --excel not provided.
2) Add per-sheet 6-parameter nondimensional correction (old->new) inside Dash:
   - Edit params in a table
   - Plot uses corrected data (optional toggle)
   - Real-time preview by re-plotting after edits
3) Export:
   - Keep Origin-format export (single X, Y+Yerr paired, blank rows between series)
   - Add extra sheets: Plotted_Long + GroupedStats (mean/std/sem/ci95)

Run:
  python figure.py
or:
  python figure.py --excel "your.xlsx" --port 8050
"""

import argparse
import os
import sys
import traceback
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime
import math

import pandas as pd
import numpy as np
import plotly.graph_objects as go

from dash import Dash, dcc, html, Input, Output, State, ctx, no_update
from dash import dash_table


# -----------------------------
# Helpers (base)
# -----------------------------
def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def sanitize_column_name(col) -> str:
    """Clean column names: remove zero-width chars, normalize spaces, strip."""
    if col is None:
        return ""
    s = str(col)
    s = s.replace("\u200b", "").replace("\ufeff", "").replace("\xa0", " ")
    s = " ".join(s.split())
    return s.strip()


def load_excel_sheets(excel_path: str) -> Tuple[Optional[Dict[str, pd.DataFrame]], List[str], Optional[str]]:
    """Load all sheets into memory. Returns: (dfs, sheet_names, error_message)."""
    if not os.path.exists(excel_path):
        return None, [], f"Excel file not found: {excel_path}"

    try:
        xls = pd.ExcelFile(excel_path)
        sheet_names = xls.sheet_names
    except Exception as ex:
        return None, [], f"Failed to open Excel: {ex}"

    dfs: Dict[str, pd.DataFrame] = {}
    for sh in sheet_names:
        try:
            df = pd.read_excel(excel_path, sheet_name=sh)
            df.columns = [sanitize_column_name(c) for c in df.columns]
            dfs[sh] = df
        except Exception as ex:
            eprint(f"[WARN] Failed to read sheet '{sh}': {ex}")
            dfs[sh] = pd.DataFrame()
    return dfs, sheet_names, None


def numeric_columns(df: pd.DataFrame) -> List[str]:
    """Return columns that can be interpreted as numeric (at least some values)."""
    cols = []
    for c in df.columns:
        if c == "" or c is None:
            continue
        s = pd.to_numeric(df[c], errors="coerce")
        if np.isfinite(s).any():
            cols.append(c)
    return cols


def safe_log_range(vmin: Optional[float], vmax: Optional[float]) -> Optional[List[float]]:
    """Convert axis min/max (linear) to Plotly log axis range (log10)."""
    if vmin is None or vmax is None:
        return None
    if not (np.isfinite(vmin) and np.isfinite(vmax)):
        return None
    if vmin <= 0 or vmax <= 0:
        return None
    if vmin == vmax:
        return None
    lo, hi = (vmin, vmax) if vmin < vmax else (vmax, vmin)
    return [np.log10(lo), np.log10(hi)]


def to_float_or_none(x) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def compute_error(y: pd.Series, mode: str) -> float:
    """
    Compute error bar size for a group.
    mode:
      - 'std': standard deviation
      - 'sem': std/sqrt(n)
      - 'ci95': ~1.96*SEM
    """
    yy = pd.to_numeric(y, errors="coerce")
    yy = yy[np.isfinite(yy)]
    n = int(yy.shape[0])
    if n <= 1:
        return 0.0
    std = float(np.nanstd(yy, ddof=1))
    if mode == "std":
        return std
    sem = std / np.sqrt(n)
    if mode == "sem":
        return float(sem)
    if mode == "ci95":
        return float(1.96 * sem)
    return std


# -----------------------------
# Non-dimensional correction (robust / partial-column)
# -----------------------------
def compute_scales_6params(rho_old: float, sigma_old: float, d_old_mm: float,
                           rho_new: float, sigma_new: float, d_new_mm: float) -> Tuple[float, float, float]:
    """Return (f_scale, trho_scale, tgamma_scale) for U unchanged."""
    d0 = float(d_old_mm) / 1000.0
    d1 = float(d_new_mm) / 1000.0
    f_scale = (float(rho_old) * d0 * d0) / (float(rho_new) * d1 * d1)
    trho_scale = d0 / d1
    tgamma_scale = math.sqrt((float(rho_old) * d0**3 / float(sigma_old)) /
                             (float(rho_new) * d1**3 / float(sigma_new)))
    return f_scale, trho_scale, tgamma_scale


def apply_nondim_correction_partial(
    df: pd.DataFrame,
    params: Dict[str, Any],
    *,
    create_missing_we: bool = False,
    append_memory_cols: bool = True,
) -> pd.DataFrame:
    """
    Apply nondimensional correction using 6 params, but only for columns that exist.
    This avoids failing when some sheets/plots don't contain the full REQUIRED_COLS set.

    Expected params keys:
      rho_old, sigma_old, d_old_mm, rho_new, sigma_new, d_new_mm  (all float)

    Notes (important for "keep Excel identical"):
    - By default we DO NOT create a missing "We" column unless create_missing_we=True.
    - By default we append memory columns (rho_target/sigma_target) for in-app visibility.
      For exporting an *identical-structure* workbook, call with append_memory_cols=False.
    """
    need = ["rho_old", "sigma_old", "d_old_mm", "rho_new", "sigma_new", "d_new_mm"]
    for k in need:
        if params.get(k, None) is None or str(params.get(k)).strip() == "":
            raise ValueError(f"Missing param: {k}")

    rho_old = float(params["rho_old"])
    sigma_old = float(params["sigma_old"])
    d_old_mm = float(params["d_old_mm"])
    rho_new = float(params["rho_new"])
    sigma_new = float(params["sigma_new"])
    d_new_mm = float(params["d_new_mm"])

    out = df.copy()
    out.columns = [sanitize_column_name(c) for c in out.columns]

    f_scale, trho_scale, tgamma_scale = compute_scales_6params(
        rho_old, sigma_old, d_old_mm, rho_new, sigma_new, d_new_mm
    )

    # Update D0 (mm) if exists
    if "D0 (mm)" in out.columns:
        out["D0 (mm)"] = float(d_new_mm)

    # Recompute We if possible (need U0)
    # IMPORTANT: only overwrite if "We" already exists, unless create_missing_we=True
    if "U0 (m/s)" in out.columns:
        u = pd.to_numeric(out["U0 (m/s)"], errors="coerce").to_numpy(dtype=float)
        d1_m = float(d_new_mm) / 1000.0
        we_new = (rho_new * (u ** 2) * d1_m) / sigma_new
        if "We" in out.columns:
            out["We"] = we_new
        elif create_missing_we:
            out["We"] = we_new

    # Scale nondimensional force columns if exist
    for c in ["F1*", "F2*"]:
        if c in out.columns:
            yy = pd.to_numeric(out[c], errors="coerce").to_numpy(dtype=float)
            out[c] = yy * f_scale

    # Scale t/τρ
    for c in ["t1/τρ", "t2/τρ"]:
        if c in out.columns:
            yy = pd.to_numeric(out[c], errors="coerce").to_numpy(dtype=float)
            out[c] = yy * trho_scale

    # Scale t/τγ
    for c in ["t1/τγ", "t2/τγ"]:
        if c in out.columns:
            yy = pd.to_numeric(out[c], errors="coerce").to_numpy(dtype=float)
            out[c] = yy * tgamma_scale

    # Optional: memory columns appended
    if append_memory_cols:
        out["rho_target"] = float(rho_new)
        out["sigma_target"] = float(sigma_new)

    return out

# -----------------------------
# Origin export helpers (existing)
# -----------------------------
def build_origin_style_table(
    series: List[Dict[str, Any]],
    x_long_name: str,
    blank_rows_between_groups: int = 2,
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Build an Origin-friendly table:

    Column layout:
      X | (Y1, Y1 Error) | (Y2, Y2 Error) | ...

    Data layout:
      - Each series occupies a contiguous block of rows (no internal blank rows)
      - Between different series blocks, add blank rows
      - Only ONE shared X column, filled with each series' X values in its own block

    Also return 'long_names' list aligned with df.columns for writing as the first row.
    """
    columns = ["X"]
    long_names = [x_long_name]

    for s in series:
        y_col = f"{s['name']}"
        e_col = f"{s['name']} Error"
        columns.extend([y_col, e_col])
        long_names.extend([s["name"], ""])

    blocks = []

    for idx, s in enumerate(series):
        x = np.asarray(s.get("x", []), dtype=float)
        y = np.asarray(s.get("y", []), dtype=float)
        yerr = s.get("yerr", None)
        if yerr is None:
            yerr = np.full_like(y, np.nan, dtype=float)
        else:
            yerr = np.asarray(yerr, dtype=float)
            if yerr.shape[0] != y.shape[0]:
                yerr = np.full_like(y, np.nan, dtype=float)

        m = np.isfinite(x) & np.isfinite(y)
        x = x[m]; y = y[m]; yerr = yerr[m]
        if x.size == 0:
            continue

        block = pd.DataFrame({c: [np.nan] * x.size for c in columns})
        block["X"] = x
        y_col = f"{s['name']}"
        e_col = f"{s['name']} Error"
        block[y_col] = y
        block[e_col] = yerr

        blocks.append(block)

        if idx != len(series) - 1 and blank_rows_between_groups > 0:
            blank = pd.DataFrame({c: [np.nan] * blank_rows_between_groups for c in columns})
            blocks.append(blank)

    if not blocks:
        return pd.DataFrame(columns=columns), long_names

    out = pd.concat(blocks, ignore_index=True)
    return out, long_names


def dataframe_to_excel_bytes_with_longname_row(df: pd.DataFrame, long_names: List[str], sheet_name: str = "OriginData") -> bytes:
    """Write an Excel file to bytes. First row = Long Name row, then data."""
    from io import BytesIO
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        long_df = pd.DataFrame([long_names], columns=df.columns)
        long_df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=0)
        df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=1)

        ws = writer.book[sheet_name]
        for i, col in enumerate(df.columns, start=1):
            # openpyxl column letters: handle >26 simply
            # Here keep it simple by using openpyxl utils
            pass

    bio.seek(0)
    return bio.read()


def export_multi_sheet_excel_bytes(
    origin_df: pd.DataFrame,
    origin_long_names: List[str],
    plotted_long_df: pd.DataFrame,
    grouped_stats_df: pd.DataFrame,
    origin_sheet: str = "OriginData",
    plotted_sheet: str = "Plotted_Long",
    stats_sheet: str = "GroupedStats",
) -> bytes:
    """Create a single xlsx (bytes) containing OriginData + Plotted_Long + GroupedStats."""
    from io import BytesIO
    from openpyxl.utils import get_column_letter

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # OriginData with Long Name row
        long_df = pd.DataFrame([origin_long_names], columns=origin_df.columns)
        long_df.to_excel(writer, index=False, header=False, sheet_name=origin_sheet, startrow=0)
        origin_df.to_excel(writer, index=False, header=False, sheet_name=origin_sheet, startrow=1)

        # Other sheets normal header
        plotted_long_df.to_excel(writer, index=False, sheet_name=plotted_sheet)
        grouped_stats_df.to_excel(writer, index=False, sheet_name=stats_sheet)

        # widen cols
        for sh in [origin_sheet, plotted_sheet, stats_sheet]:
            ws = writer.book[sh]
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                letter = get_column_letter(col_idx)
                ws.column_dimensions[letter].width = 16

    bio.seek(0)
    return bio.read()


# -----------------------------
# Export: update ORIGINAL workbook values (keep sheets & formatting)
# -----------------------------
def _copy_cell_style(src_cell, dst_cell):
    """Copy openpyxl cell style to keep workbook formatting consistent."""
    try:
        from copy import copy
        dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)
    except Exception:
        pass


def _find_first_match(cols: List[str], candidates: List[str]) -> Optional[str]:
    low = [c.lower() for c in cols]
    for cand in candidates:
        cl = cand.lower()
        if cl in low:
            return cols[low.index(cl)]
    return None


def _guess_liquid_col(df: pd.DataFrame) -> Optional[str]:
    """Locate the 'liquid/solution' column (best effort)."""
    cols = [sanitize_column_name(c) for c in df.columns]
    candidates = [
        "Liquid", "Fluid", "Solution", "Solvent", "Sample", "Material",
        "液体", "溶液", "溶剂", "样品", "材料",
    ]
    return _find_first_match(cols, candidates)


def _extract_concentration(text: Any) -> Optional[str]:
    """Best-effort concentration parser from a string like 'Water 10%', 'NaCl 0.1M'."""
    if text is None:
        return None
    s = str(text).strip()
    if not s:
        return None
    pats = [
        r"(\d+(?:\.\d+)?)\s*(wt%|w/w|v/v|vol%|%|mM|M|mol/L|mol·L-1|g/L|mg/mL|mg/L)",
        r"(\d+(?:\.\d+)?)\s*(?:%)",
    ]
    import re as _re
    for p in pats:
        mm = _re.search(p, s, flags=_re.I)
        if mm:
            return "".join(mm.groups())
    return None


def _ensure_physical_cols_in_df(
    df: pd.DataFrame,
    *,
    liquid_col: Optional[str],
    rho_value: Optional[float],
    sigma_value: Optional[float],
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Ensure these columns exist:
      - Density (kg/m3)
      - SurfaceTension (N/m)
      - Concentration
    """
    out = df.copy()
    out.columns = [sanitize_column_name(c) for c in out.columns]
    new_cols = []

    col_rho = "Density (kg/m3)"
    col_sigma = "SurfaceTension (N/m)"
    col_conc = "Concentration"

    if col_rho not in out.columns:
        out[col_rho] = np.nan
        new_cols.append(col_rho)
    if col_sigma not in out.columns:
        out[col_sigma] = np.nan
        new_cols.append(col_sigma)
    if col_conc not in out.columns:
        out[col_conc] = ""
        new_cols.append(col_conc)

    if rho_value is not None:
        out[col_rho] = float(rho_value)
    if sigma_value is not None:
        out[col_sigma] = float(sigma_value)

    if liquid_col and liquid_col in out.columns:
        conc = out[liquid_col].apply(_extract_concentration)
        out.loc[out[col_conc].astype(str).str.strip().eq(""), col_conc] = conc.fillna("")

    return out, new_cols


def _update_sheet_values_keep_format(
    ws,
    df_new: pd.DataFrame,
    *,
    header_row: int = 1,
    insert_after_col_name: Optional[str] = None,
    insert_cols_order: Optional[List[str]] = None,
) -> List[str]:
    """
    Update values in-place on an openpyxl worksheet:
    - Keep formatting (we only write cell.value).
    - Optionally insert missing physical columns after a specified anchor column.
    """
    warnings = []

    header_cells = list(ws[header_row])
    header_vals = [sanitize_column_name(c.value) for c in header_cells]
    col_to_idx = {name: i + 1 for i, name in enumerate(header_vals) if name}

    # Insert missing columns (only those listed)
    if insert_cols_order:
        after_name = sanitize_column_name(insert_after_col_name) if insert_after_col_name else ""
        after_idx = col_to_idx.get(after_name, None) if after_name else None
        if after_idx is None:
            if insert_after_col_name:
                warnings.append(f"Insert anchor column not found: {after_name}. Physical cols appended to end.")
            after_idx = ws.max_column

        insert_pos = after_idx + 1
        for col_name in insert_cols_order:
            if col_name in col_to_idx:
                continue

            ws.insert_cols(insert_pos, 1)

            # header value + style
            src_h = ws.cell(row=header_row, column=after_idx)
            dst_h = ws.cell(row=header_row, column=insert_pos)
            dst_h.value = col_name
            _copy_cell_style(src_h, dst_h)

            # copy style down this new column from anchor column
            max_row = max(ws.max_row, header_row + 1 + int(df_new.shape[0]))
            for r in range(header_row + 1, max_row + 1):
                _copy_cell_style(ws.cell(row=r, column=after_idx), ws.cell(row=r, column=insert_pos))

            # rebuild mapping (because columns shifted)
            header_cells = list(ws[header_row])
            header_vals = [sanitize_column_name(c.value) for c in header_cells]
            col_to_idx = {name: i + 1 for i, name in enumerate(header_vals) if name}

            insert_pos += 1

    df_new = df_new.copy()
    df_new.columns = [sanitize_column_name(c) for c in df_new.columns]

    n = int(df_new.shape[0])
    start_row = header_row + 1
    end_row = start_row + n - 1

    for col_name in df_new.columns:
        if col_name not in col_to_idx:
            continue
        col_idx = col_to_idx[col_name]
        vals = df_new[col_name].tolist()

        for i, v in enumerate(vals):
            r = start_row + i
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                cell.value = None
            elif pd.isna(v):
                cell.value = None
            else:
                if isinstance(v, (np.generic,)):
                    v = v.item()
                cell.value = v

        # If sheet has more rows, clear remaining values in the updated columns
        if ws.max_row > end_row:
            for r in range(end_row + 1, ws.max_row + 1):
                c = ws.cell(row=r, column=col_idx)
                if c.value is not None:
                    c.value = None

    return warnings


def export_corrected_workbook_bytes(
    excel_path: str,
    dfs: Dict[str, pd.DataFrame],
    selected_sheets: List[str],
    *,
    use_nondim: bool,
    params_by_sheet: Dict[str, Dict[str, Any]],
) -> Tuple[bytes, List[str]]:
    """
    Export a workbook that is identical in structure & formatting,
    but updates the DATA for the selected sheets.

    Also inserts three physical columns right after the Liquid column (if found):
      Density (kg/m3), SurfaceTension (N/m), Concentration
    """
    warnings: List[str] = []
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)

    for sh in selected_sheets:
        if sh not in wb.sheetnames:
            warnings.append(f"[{sh}] Sheet not found in workbook; skipped.")
            continue

        ws = wb[sh]
        df0 = dfs.get(sh, pd.DataFrame())
        if df0 is None or df0.empty:
            warnings.append(f"[{sh}] Sheet empty in pandas; skipped.")
            continue

        df_new = df0.copy()

        rho_val, sigma_val = None, None
        if use_nondim:
            p = params_by_sheet.get(sh, None)
            if p is None:
                warnings.append(f"[{sh}] Non-dim ON but no params row; export uses raw values.")
            else:
                try:
                    df_new = apply_nondim_correction_partial(
                        df0,
                        p,
                        create_missing_we=False,   # keep structure identical
                        append_memory_cols=False,  # keep structure identical
                    )
                    rv = str(p.get("rho_new", "")).strip()
                    sv = str(p.get("sigma_new", "")).strip()
                    rho_val = float(rv) if rv else None
                    sigma_val = float(sv) if sv else None
                except Exception as ex:
                    warnings.append(f"[{sh}] Non-dim failed: {ex}; export uses raw values.")
                    df_new = df0.copy()

        liquid_col = _guess_liquid_col(df_new)
        df_new2, _ = _ensure_physical_cols_in_df(
            df_new,
            liquid_col=liquid_col,
            rho_value=rho_val,
            sigma_value=sigma_val,
        )

        insert_cols_order = ["Density (kg/m3)", "SurfaceTension (N/m)", "Concentration"]

        warns2 = _update_sheet_values_keep_format(
            ws,
            df_new2,
            header_row=1,
            insert_after_col_name=liquid_col,
            insert_cols_order=insert_cols_order,
        )
        warnings.extend([f"[{sh}] {w}" for w in warns2])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read(), warnings

# -----------------------------
# Dash App
# -----------------------------
def build_app(excel_path: str) -> Dash:
    dfs, sheet_names, err = load_excel_sheets(excel_path)

    app = Dash(__name__)
    app.title = "Excel Multi-Sheet Log-Log Plotter"

    if err:
        app.layout = html.Div(
            style={"fontFamily": "Arial, sans-serif", "padding": "16px"},
            children=[
                html.H2("Excel Load Error"),
                html.Pre(err, style={"whiteSpace": "pre-wrap", "color": "crimson"}),
                html.P("Please check the file path and try again."),
            ],
        )
        return app

    # Union of numeric columns across sheets
    all_numeric_cols = set()
    for sh in sheet_names:
        df = dfs.get(sh, pd.DataFrame())
        for c in numeric_columns(df):
            all_numeric_cols.add(c)
    all_numeric_cols = sorted(all_numeric_cols)

    def pick_default(col_candidates: List[str], preferred: List[str]) -> Optional[str]:
        low = [c.lower() for c in col_candidates]
        for p in preferred:
            if p.lower() in low:
                return col_candidates[low.index(p.lower())]
        return col_candidates[0] if col_candidates else None

    default_x = pick_default(all_numeric_cols, ["We", "we", "Weber", "Weber number"])
    default_y = pick_default(all_numeric_cols, ["F1*", "F2*", "t1/τρ", "t1/τγ", "t2/τρ", "t2/τγ"])

    # Build default param table rows
    def guess_d0_mm(df: pd.DataFrame) -> Optional[float]:
        if "D0 (mm)" not in df.columns:
            return None
        s = pd.to_numeric(df["D0 (mm)"], errors="coerce")
        u = s[np.isfinite(s)].unique()
        if u.size == 1:
            return float(u[0])
        return None

    def guess_targets(df: pd.DataFrame, col: str) -> Optional[float]:
        if col not in df.columns:
            return None
        s = pd.to_numeric(df[col], errors="coerce")
        u = s[np.isfinite(s)].unique()
        if u.size == 1:
            return float(u[0])
        return None

    param_rows = []
    for sh in sheet_names:
        df = dfs.get(sh, pd.DataFrame())
        d0_guess = guess_d0_mm(df)
        rho_guess = guess_targets(df, "rho_target")
        sig_guess = guess_targets(df, "sigma_target")
        param_rows.append({
            "sheet": sh,
            "rho_old": "",
            "sigma_old": "",
            "d_old_mm": "" if d0_guess is None else str(d0_guess),
            "rho_new": "" if rho_guess is None else str(rho_guess),
            "sigma_new": "" if sig_guess is None else str(sig_guess),
            "d_new_mm": "" if d0_guess is None else str(d0_guess),
        })

    app.layout = html.Div(
        style={"fontFamily": "Arial, sans-serif", "padding": "14px"},
        children=[
            html.H2("Excel Multi-Sheet Log-Log Plotter (Plotly + Dash) + Non-dimensional Correction"),

            html.Div(
                style={"padding": "10px", "border": "1px solid #ddd", "borderRadius": "8px", "marginBottom": "12px"},
                children=[
                    html.Div(
                        style={"display": "flex", "gap": "12px", "flexWrap": "wrap"},
                        children=[
                            html.Div(
                                style={"minWidth": "280px", "flex": "1"},
                                children=[
                                    html.Label("Worksheets (multi-select)"),
                                    dcc.Dropdown(
                                        id="sheets",
                                        options=[{"label": s, "value": s} for s in sheet_names],
                                        value=[sheet_names[0]] if sheet_names else [],
                                        multi=True,
                                        placeholder="Select one or more sheets...",
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "240px", "flex": "1"},
                                children=[
                                    html.Label("X column (log axis)"),
                                    dcc.Dropdown(
                                        id="xcol",
                                        options=[{"label": c, "value": c} for c in all_numeric_cols],
                                        value=default_x,
                                        clearable=False,
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "240px", "flex": "1"},
                                children=[
                                    html.Label("Y column (log axis)"),
                                    dcc.Dropdown(
                                        id="ycol",
                                        options=[{"label": c, "value": c} for c in all_numeric_cols],
                                        value=default_y,
                                        clearable=False,
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "220px"},
                                children=[
                                    html.Label("Trace style (raw)"),
                                    dcc.Dropdown(
                                        id="mode",
                                        options=[
                                            {"label": "Markers", "value": "markers"},
                                            {"label": "Lines", "value": "lines"},
                                            {"label": "Lines + Markers", "value": "lines+markers"},
                                        ],
                                        value="markers",
                                        clearable=False,
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "180px"},
                                children=[
                                    html.Label("Legend"),
                                    dcc.Dropdown(
                                        id="legend",
                                        options=[{"label": "Show", "value": "show"}, {"label": "Hide", "value": "hide"}],
                                        value="show",
                                        clearable=False,
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "260px"},
                                children=[
                                    html.Label("Non-dimensional correction"),
                                    dcc.Dropdown(
                                        id="use_nondim",
                                        options=[
                                            {"label": "Off (use raw Excel values)", "value": "off"},
                                            {"label": "On (apply per-sheet 6-params)", "value": "on"},
                                        ],
                                        value="off",
                                        clearable=False,
                                    ),
                                ],
                            ),
                        ],
                    ),
                    html.Hr(),

                    html.Div(
                        style={"display": "flex", "gap": "12px", "flexWrap": "wrap"},
                        children=[
                            html.Div(
                                style={"minWidth": "320px", "flex": "1"},
                                children=[
                                    html.Label("X range (linear values, for log axis). Leave blank for auto."),
                                    html.Div(
                                        style={"display": "flex", "gap": "8px"},
                                        children=[
                                            dcc.Input(id="xmin", type="text", placeholder="xmin (>0)", style={"width": "50%"}),
                                            dcc.Input(id="xmax", type="text", placeholder="xmax (>0)", style={"width": "50%"}),
                                        ],
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "320px", "flex": "1"},
                                children=[
                                    html.Label("Y range (linear values, for log axis). Leave blank for auto."),
                                    html.Div(
                                        style={"display": "flex", "gap": "8px"},
                                        children=[
                                            dcc.Input(id="ymin", type="text", placeholder="ymin (>0)", style={"width": "50%"}),
                                            dcc.Input(id="ymax", type="text", placeholder="ymax (>0)", style={"width": "50%"}),
                                        ],
                                    ),
                                ],
                            ),
                            html.Div(
                                style={"minWidth": "220px"},
                                children=[
                                    html.Label("Point size"),
                                    dcc.Slider(id="psize", min=3, max=14, step=1, value=7, marks={3: "3", 7: "7", 14: "14"}),
                                ],
                            ),
                        ],
                    ),

                    html.Div(
                        style={"marginTop": "10px", "display": "flex", "gap": "10px", "alignItems": "center", "flexWrap": "wrap"},
                        children=[
                            html.Button("Plot raw", id="plot_raw", n_clicks=0),
                            html.Button("Mean + error bars (grouped by same X)", id="plot_mean", n_clicks=0),
                            html.Div(
                                style={"minWidth": "240px"},
                                children=[
                                    html.Label("Error bar type (for mean plot & stats export)"),
                                    dcc.Dropdown(
                                        id="err_type",
                                        options=[
                                            {"label": "STD (standard deviation)", "value": "std"},
                                            {"label": "SEM (std/sqrt(n))", "value": "sem"},
                                            {"label": "95% CI (~1.96*SEM)", "value": "ci95"},
                                        ],
                                        value="std",
                                        clearable=False,
                                    ),
                                ],
                            ),
                            html.Div(style={"minWidth": "420px"}, children=[
                                html.Label("Export"),
                                html.Button("Export plotted data (Origin + Stats)", id="export_btn", n_clicks=0),
                                html.Button("Export corrected workbook (same sheets)", id="export_corrected_btn", n_clicks=0,
                                            style={"marginLeft": "10px"}),
                            ]),
                            html.Span("Tip: Edit params below → switch correction ON → re-plot to preview.", style={"color": "#444"}),
                        ],
                    ),
                ],
            ),

            html.Div(style={"marginBottom": "10px"}, children=[
                html.Div(id="warnings", style={"whiteSpace": "pre-wrap", "color": "crimson"}),
            ]),

            html.H4("Per-sheet 6-parameter table (edit here; used when correction is ON)"),
            dash_table.DataTable(
                id="param_table",
                columns=[
                    {"name": "sheet", "id": "sheet", "editable": False},
                    {"name": "rho_old", "id": "rho_old", "type": "text"},
                    {"name": "sigma_old", "id": "sigma_old", "type": "text"},
                    {"name": "d_old_mm", "id": "d_old_mm", "type": "text"},
                    {"name": "rho_new", "id": "rho_new", "type": "text"},
                    {"name": "sigma_new", "id": "sigma_new", "type": "text"},
                    {"name": "d_new_mm", "id": "d_new_mm", "type": "text"},
                ],
                data=param_rows,
                editable=True,
                page_size=12,
                style_table={"overflowX": "auto"},
                style_cell={"minWidth": "110px", "width": "110px", "maxWidth": "160px", "whiteSpace": "normal"},
                style_header={"fontWeight": "bold"},
            ),

            dcc.Loading(
                type="default",
                children=[
                    dcc.Graph(
                        id="fig",
                        style={"height": "70vh"},
                        config={"displayModeBar": True, "scrollZoom": True, "toImageButtonOptions": {"format": "png", "filename": "export_plot"}},
                    )
                ],
            ),

            # Store the "currently plotted data" so export can output exactly what's shown
            dcc.Store(id="plotted_store", data={"mode": "raw", "xcol": "", "ycol": "", "series": [], "use_nondim": "off"}),

            dcc.Download(id="download_xlsx"),

            dcc.Download(id="download_corrected_xlsx"),
        ],
    )

    def params_table_to_dict(table_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        """Convert DataTable rows to {sheet: params_dict}"""
        out = {}
        for r in table_rows or []:
            sh = str(r.get("sheet", "")).strip()
            if not sh:
                continue
            out[sh] = {
                "rho_old": r.get("rho_old", ""),
                "sigma_old": r.get("sigma_old", ""),
                "d_old_mm": r.get("d_old_mm", ""),
                "rho_new": r.get("rho_new", ""),
                "sigma_new": r.get("sigma_new", ""),
                "d_new_mm": r.get("d_new_mm", ""),
            }
        return out

    # ---------- Plot callback (updates figure + store) ----------
    @app.callback(
        Output("fig", "figure"),
        Output("warnings", "children"),
        Output("plotted_store", "data"),
        Input("plot_raw", "n_clicks"),
        Input("plot_mean", "n_clicks"),
        State("sheets", "value"),
        State("xcol", "value"),
        State("ycol", "value"),
        State("mode", "value"),
        State("legend", "value"),
        State("xmin", "value"),
        State("xmax", "value"),
        State("ymin", "value"),
        State("ymax", "value"),
        State("psize", "value"),
        State("err_type", "value"),
        State("use_nondim", "value"),
        State("param_table", "data"),
    )
    def update_figure(_n_raw, _n_mean, sheets, xcol, ycol, mode, legend,
                      xmin, xmax, ymin, ymax, psize, err_type, use_nondim, param_table_data):
        warn_lines = []
        triggered = ctx.triggered_id
        do_mean = (triggered == "plot_mean")

        store_payload = {
            "mode": "mean" if do_mean else "raw",
            "xcol": xcol or "",
            "ycol": ycol or "",
            "series": [],
            "use_nondim": use_nondim or "off",
        }

        if not sheets:
            fig = go.Figure()
            fig.update_layout(title="No worksheet selected",
                              xaxis={"type": "log", "title": xcol or "X"},
                              yaxis={"type": "log", "title": ycol or "Y"})
            return fig, "Please select at least one worksheet.", store_payload

        if xcol is None or ycol is None:
            fig = go.Figure()
            fig.update_layout(title="Missing axis columns")
            return fig, "Please select both X and Y columns.", store_payload

        params_by_sheet = params_table_to_dict(param_table_data)

        fig = go.Figure()
        total_dropped = 0
        total_points = 0
        total_groups = 0

        for sh in sheets:
            df0 = dfs.get(sh, pd.DataFrame())
            if df0 is None or df0.empty:
                warn_lines.append(f"[{sh}] Sheet is empty or failed to read.")
                continue

            df = df0

            # apply nondim correction if ON
            if use_nondim == "on":
                p = params_by_sheet.get(sh, None)
                if p is None:
                    warn_lines.append(f"[{sh}] Non-dim is ON but no params row found.")
                else:
                    try:
                        df = apply_nondim_correction_partial(df0, p)
                    except Exception as ex:
                        warn_lines.append(f"[{sh}] Non-dim params invalid or failed: {ex}")
                        df = df0  # fallback to raw

            if xcol not in df.columns or ycol not in df.columns:
                warn_lines.append(f"[{sh}] Missing columns: {xcol} or {ycol}")
                continue

            x = pd.to_numeric(df[xcol], errors="coerce")
            y = pd.to_numeric(df[ycol], errors="coerce")

            valid = np.isfinite(x) & np.isfinite(y)
            log_valid = valid & (x > 0) & (y > 0)

            dropped = int(valid.sum() - log_valid.sum())
            total_dropped += max(dropped, 0)

            if log_valid.sum() == 0:
                warn_lines.append(f"[{sh}] No valid points after filtering (need X>0 and Y>0).")
                continue

            x_valid = x[log_valid].to_numpy(dtype=float)
            y_valid = y[log_valid].to_numpy(dtype=float)

            if not do_mean:
                total_points += int(x_valid.shape[0])
                trace_name = sh + (" [ND]" if use_nondim == "on" else "")
                fig.add_trace(go.Scatter(x=x_valid, y=y_valid, mode=mode, name=trace_name, marker={"size": int(psize)}))
                store_payload["series"].append({"name": trace_name, "x": x_valid.tolist(), "y": y_valid.tolist(), "yerr": None, "n": None})

            else:
                tmp = pd.DataFrame({"x": x_valid, "y": y_valid})
                g = tmp.groupby("x", sort=True)
                x_mean, y_mean, y_err, n_list = [], [], [], []

                for xv, sub in g:
                    yy = sub["y"]
                    m = float(np.nanmean(yy))
                    if not np.isfinite(m) or m <= 0:
                        continue
                    errv = compute_error(yy, err_type)
                    x_mean.append(float(xv))
                    y_mean.append(m)
                    y_err.append(float(errv))
                    n_list.append(int(sub.shape[0]))

                if len(x_mean) == 0:
                    warn_lines.append(f"[{sh}] No valid groups after averaging (check Y>0 and grouping).")
                    continue

                total_groups += len(x_mean)
                bad_lower = sum(1 for m, e in zip(y_mean, y_err) if (m - e) <= 0 and e > 0)
                if bad_lower > 0:
                    warn_lines.append(f"[{sh}] {bad_lower} points have (mean - error) <= 0 on log axis; consider SEM/CI or check data.")

                trace_name = f"{sh} (mean±{err_type})" + (" [ND]" if use_nondim == "on" else "")
                fig.add_trace(
                    go.Scatter(
                        x=x_mean,
                        y=y_mean,
                        mode="markers+lines",
                        name=trace_name,
                        marker={"size": int(psize) + 2},
                        error_y={"type": "data", "array": y_err, "visible": True},
                        hovertemplate="x=%{x}<br>mean=%{y}<br>error=%{customdata}<br>n=%{text}<extra></extra>",
                        customdata=y_err,
                        text=n_list,
                    )
                )
                store_payload["series"].append({"name": trace_name, "x": x_mean, "y": y_mean, "yerr": y_err, "n": n_list})

        xmin_f, xmax_f = to_float_or_none(xmin), to_float_or_none(xmax)
        ymin_f, ymax_f = to_float_or_none(ymin), to_float_or_none(ymax)
        xr = safe_log_range(xmin_f, xmax_f)
        yr = safe_log_range(ymin_f, ymax_f)

        if (xmin_f is not None or xmax_f is not None) and xr is None:
            warn_lines.append("[X range] Invalid range. Please provide xmin and xmax as numbers > 0 and not equal.")
        if (ymin_f is not None or ymax_f is not None) and yr is None:
            warn_lines.append("[Y range] Invalid range. Please provide ymin and ymax as numbers > 0 and not equal.")

        if total_dropped > 0:
            warn_lines.append(f"Dropped {total_dropped} rows with non-positive values (required for log axes).")

        if not do_mean and total_points == 0:
            warn_lines.append("No points to plot (all selected sheets empty/invalid, or filtered by log constraint).")
        if do_mean and total_groups == 0:
            warn_lines.append("No averaged groups to plot (all selected sheets empty/invalid, or filtered by log constraint).")

        title_suffix = "Mean + error bars (grouped by same X)" if do_mean else "Raw"
        nd_suffix = " + Non-dimensional" if use_nondim == "on" else ""
        fig.update_layout(
            title=f"Log-Log Plot ({title_suffix}{nd_suffix}): {ycol} vs {xcol}",
            xaxis={"type": "log", "title": xcol, "range": xr, "showgrid": True, "zeroline": False},
            yaxis={"type": "log", "title": ycol, "range": yr, "showgrid": True, "zeroline": False},
            showlegend=(legend == "show"),
            margin={"l": 60, "r": 30, "t": 60, "b": 70},
            hovermode="closest",
        )
        fig.update_layout(
            annotations=[dict(
                text="Tip: Edit params table → set correction ON → click Plot to preview.",
                x=0, y=-0.18, xref="paper", yref="paper", showarrow=False, font={"size": 12},
            )]
        )

        return fig, "\n".join(warn_lines), store_payload

    # ---------- Export callback (uses plotted_store) ----------
    @app.callback(
        Output("download_xlsx", "data"),
        Output("warnings", "children", allow_duplicate=True),
        Input("export_btn", "n_clicks"),
        State("plotted_store", "data"),
        State("err_type", "value"),
        prevent_initial_call=True,
    )
    def export_current_plotted_data(_n_clicks, plotted_store, err_type):
        if not plotted_store or not plotted_store.get("series"):
            return no_update, "No plotted data to export yet. Please plot first."

        try:
            xcol = plotted_store.get("xcol", "X")
            ycol = plotted_store.get("ycol", "Y")
            mode = plotted_store.get("mode", "raw")
            series = plotted_store.get("series", [])
            use_nondim = plotted_store.get("use_nondim", "off")

            # 1) Origin-format table (what you already had)
            origin_df, long_names = build_origin_style_table(
                series=series,
                x_long_name=xcol if xcol else "X",
                blank_rows_between_groups=2,
            )

            # 2) Plotted long table (each point)
            long_rows = []
            for s in series:
                name = s.get("name", "")
                xs = s.get("x", []) or []
                ys = s.get("y", []) or []
                yerrs = s.get("yerr", None)
                ns = s.get("n", None)

                if yerrs is None:
                    yerrs = [np.nan] * len(xs)
                if ns is None:
                    ns = [np.nan] * len(xs)

                for xv, yv, ev, nv in zip(xs, ys, yerrs, ns):
                    long_rows.append({
                        "trace": name,
                        "x": xv,
                        "y": yv,
                        "yerr": ev,
                        "n": nv,
                        "xcol": xcol,
                        "ycol": ycol,
                        "mode": mode,
                        "nondim": use_nondim,
                    })
            plotted_long_df = pd.DataFrame(long_rows)

            # 3) Grouped stats (mainly for raw mode; for mean mode it mirrors)
            stats_rows = []
            if mode == "raw":
                # compute mean/std/sem/ci95 per trace per x
                if not plotted_long_df.empty:
                    for trace_name, g0 in plotted_long_df.groupby("trace"):
                        g0 = g0.dropna(subset=["x", "y"])
                        if g0.empty:
                            continue
                        for xv, g1 in g0.groupby("x"):
                            yvals = pd.to_numeric(g1["y"], errors="coerce").to_numpy(dtype=float)
                            yvals = yvals[np.isfinite(yvals)]
                            if yvals.size == 0:
                                continue
                            n = int(yvals.size)
                            mean = float(np.nanmean(yvals))
                            std = float(np.nanstd(yvals, ddof=1)) if n > 1 else 0.0
                            sem = float(std / np.sqrt(n)) if n > 1 else 0.0
                            ci95 = float(1.96 * sem) if n > 1 else 0.0
                            stats_rows.append({
                                "trace": trace_name,
                                "x": float(xv),
                                "n": n,
                                "mean": mean,
                                "std": std,
                                "sem": sem,
                                "ci95": ci95,
                                "err_type_selected": err_type,
                            })
            else:
                # mean mode: use stored yerr as "selected error"
                if not plotted_long_df.empty:
                    for trace_name, g0 in plotted_long_df.groupby("trace"):
                        g0 = g0.dropna(subset=["x", "y"])
                        for _, r in g0.iterrows():
                            stats_rows.append({
                                "trace": trace_name,
                                "x": float(r["x"]),
                                "n": r.get("n", np.nan),
                                "mean": float(r["y"]),
                                "selected_error": float(r["yerr"]) if pd.notna(r["yerr"]) else np.nan,
                                "err_type_selected": err_type,
                            })

            grouped_stats_df = pd.DataFrame(stats_rows)

            # filename
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            ndtag = "ND" if use_nondim == "on" else "RAW"
            filename = f"Export_{ndtag}_{xcol}_vs_{ycol}_{mode}_{ts}.xlsx"
            for bad in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
                filename = filename.replace(bad, "_")

            xlsx_bytes = export_multi_sheet_excel_bytes(
                origin_df=origin_df,
                origin_long_names=long_names,
                plotted_long_df=plotted_long_df if not plotted_long_df.empty else pd.DataFrame(columns=["trace","x","y","yerr","n","xcol","ycol","mode","nondim"]),
                grouped_stats_df=grouped_stats_df if not grouped_stats_df.empty else pd.DataFrame(),
            )

            return dcc.send_bytes(xlsx_bytes, filename), no_update

        except Exception as ex:
            err_msg = "[Export ERROR] Failed to export Excel:\n" + str(ex)
            eprint(err_msg)
            eprint(traceback.format_exc())
            return no_update, err_msg


    # ---------- Export callback: corrected workbook (same sheets & formatting) ----------
    @app.callback(
        Output("download_corrected_xlsx", "data"),
        Output("warnings", "children", allow_duplicate=True),
        Input("export_corrected_btn", "n_clicks"),
        State("sheets", "value"),
        State("use_nondim", "value"),
        State("param_table", "data"),
        prevent_initial_call=True,
    )
    def export_corrected_workbook(_n_clicks, sheets, use_nondim, param_table_data):
        if not sheets:
            return no_update, "No worksheet selected for corrected export."

        try:
            params_by_sheet = params_table_to_dict(param_table_data)
            xbytes, warns = export_corrected_workbook_bytes(
                excel_path=excel_path,
                dfs=dfs,
                selected_sheets=sheets,
                use_nondim=(use_nondim == "on"),
                params_by_sheet=params_by_sheet,
            )

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"CorrectedWorkbook_{ts}.xlsx"
            warn_text = "\n".join(warns) if warns else no_update
            return dcc.send_bytes(xbytes, filename), warn_text

        except Exception as ex:
            err_msg = "[Corrected Export ERROR] Failed to export corrected workbook:\n" + str(ex)
            eprint(err_msg)
            eprint(traceback.format_exc())
            return no_update, err_msg


    return app


def prompt_excel_path_if_missing(p: Optional[str]) -> str:
    if p and str(p).strip():
        return str(p).strip().strip('"').strip("'")
    # runtime prompt
    print("\n[Input required] Please enter the Excel file path (.xlsx):")
    p2 = input("Excel path: ").strip().strip('"').strip("'")
    return p2


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=str, default="", help="Excel path (.xlsx). If empty, you will be prompted at runtime.")
    parser.add_argument("--host", type=str, default="127.0.0.1", help="Host to bind. Default: 127.0.0.1")
    parser.add_argument("--port", type=int, default=8050, help="Port to bind. Default: 8050")
    parser.add_argument("--debug", action="store_true", help="Enable Dash debug mode")
    args = parser.parse_args()

    excel_path = prompt_excel_path_if_missing(args.excel)
    if not excel_path:
        eprint("[ERROR] No Excel path provided.")
        sys.exit(1)

    # keep relative path allowed: relative to current working directory
    excel_path = os.path.abspath(excel_path)

    if not os.path.exists(excel_path):
        eprint(f"[ERROR] Excel file not found: {excel_path}")
        sys.exit(1)

    app = build_app(excel_path)

    try:
        app.run(host=args.host, port=args.port, debug=args.debug)
    except OSError as ex:
        eprint(f"[ERROR] Failed to start server: {ex}")
        eprint("Try a different port: python figure.py --port 8051")
    except Exception:
        eprint("[ERROR] Unexpected error:")
        eprint(traceback.format_exc())


if __name__ == "__main__":
    main()
